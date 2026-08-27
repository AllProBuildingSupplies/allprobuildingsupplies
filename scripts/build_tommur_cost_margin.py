#!/usr/bin/env python3
"""
Build Tommur cost/margin tracker workbook by merging:
- All Current Supplies (Tommur-offerable catalog)
- All 3 Projects / Margins (yellow FOB + DDP + carton dims)
- Davenport PI (carton dims / alternate FOB)
- Lesso DWV list (Lesso codes + carton dims)
- Website products.csv (APBS item codes + selling prices)

Fittings customer price list was NOT available in this environment
(path was local OneDrive). Website prices used where present; blanks otherwise.
"""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

UPLOADS = Path("/home/ubuntu/.cursor/projects/workspace/uploads")
PRODUCTS_CSV = Path("/workspace/assets/products.csv")
OUT_PATH = Path("/workspace/costing/Tommur_Cost_Margin_Tracker.xlsx")

# Container internal volumes (CBM) — industry standard averages
CBM_40FT = 67.7  # standard 40'
CBM_45HQ = 86.0  # 45' high cube
FREIGHT_PER_CONTAINER = 7000.0  # USD

# Tariff stack as of ~Aug 2026 research (verify with broker before entry):
# MFN base + Section 301 China (+25% typical for these headings)
# + additional China FLIP/forced-labor Section 301 (+12.5% from Jul 24, 2026)
ADD_SEC301 = 0.25
ADD_FLIP = 0.125

HTS_TABLE = [
    # category, hts, mfn, notes
    (
        "PVC_DWV_FITTING",
        "3917.40.00.20",
        0.053,
        "Plumbing fittings, not pressure rated (DWV), of PVC",
    ),
    (
        "PVC_PRESSURE_FITTING",
        "3917.40.00.60",
        0.053,
        "Other plumbing fittings of PVC (pressure-rated / SCH)",
    ),
    (
        "PVC_PIPE_RIGID",
        "3917.23.00.00",
        0.031,
        "Rigid tubes/pipes of polymers of vinyl chloride (SCH40/SCH80/foam-core PVC)",
    ),
    (
        "CPVC_PIPE_RIGID",
        "3917.23.00.00",
        0.031,
        "Rigid tubes/pipes of polymers of vinyl chloride (CPVC SCH80 / SDR)",
    ),
    (
        "CPVC_FITTING",
        "3917.40.00.60",
        0.053,
        "Other plumbing fittings of PVC/CPVC (pressure)",
    ),
    (
        "PEX_PIPE",
        "3917.21.00.00",
        0.031,
        "Tubes/pipes of polymers of ethylene (PEX-B)",
    ),
    (
        "PEX_FITTING",
        "3917.40.00.80",
        0.053,
        "Other plastic plumbing fittings (PEX elbows/reducers etc.)",
    ),
    (
        "COPPER_PIPE",
        "7411.10.10.90",
        0.015,
        "Seamless tubes of refined copper (Type K/L pipe) — confirm OD/coil vs straight",
    ),
    (
        "COPPER_FITTING",
        "7412.10.00.00",
        0.03,
        "Copper tube/pipe fittings of refined copper — brass alloy fittings may be 7412.20",
    ),
    (
        "INSULATION",
        "3926.90.99.85",
        0.053,
        "Other articles of plastics (pipe insulation) — confirm with broker; alt 3917/3921",
    ),
]


def norm_text(s) -> str:
    if s is None:
        return ""
    s = str(s).replace("\xa0", " ").replace("\n", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_name(s) -> str:
    s = norm_text(s).upper()
    s = s.replace("×", "X").replace("ｘ", "X")
    s = s.replace("(H × H)", "(H X H)").replace("(H × S)", "(H X S)")
    s = s.replace("(S × H)", "(S X H)").replace("(H × FPT)", "(H X FPT)")
    s = re.sub(r"[‘’′']", "", s)
    s = re.sub(r"\s+", " ", s)
    # common website vs factory wording
    s = s.replace("(H X H)", "(H X H)")
    s = s.replace("1/16 BEND (H X H)", "1/16 BEND (H X H)")
    s = s.replace("ELBOW 90", "ELBOW 90")
    return s.strip()


def norm_size(s) -> str:
    if s is None:
        return ""
    s = norm_text(s)
    s = s.replace("×", "x").replace("Ｘ", "x").replace("X", "x")
    s = re.sub(r'[″"”]', '"', s)
    s = re.sub(r"[′']", "", s)
    # drop metric annotations
    s = re.sub(r"\([^)]*mm[^)]*\)", "", s, flags=re.I)
    s = re.sub(r"（[^）]*）", "", s)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"\s*x\s*", "x", s)
    s = re.sub(r"\s+", "", s)
    s = s.replace("−", "-").replace("–", "-")
    # normalize fractions spacing already removed
    s = s.replace("11/2", "1-1/2").replace("11/4", "1-1/4")
    s = s.replace("21/2", "2-1/2")
    # unify 1 1/2 -> 1-1/2 if still present
    s = re.sub(r"(\d)(\d/\d)", r"\1-\2", s)
    if s and not s.endswith('"') and re.search(r"\d", s):
        s = s + '"'
    return s


def parse_carton_cm(val):
    """Parse '56×42×33' or similar into L,W,H floats."""
    if val is None:
        return None, None, None
    if isinstance(val, (int, float)):
        return None, None, None
    s = norm_text(val).lower().replace("×", "x").replace("*", "x")
    m = re.findall(r"(\d+(?:\.\d+)?)", s)
    if len(m) >= 3:
        return float(m[0]), float(m[1]), float(m[2])
    return None, None, None


def cbm_from_lwh(l, w, h):
    if l and w and h and l > 0 and w > 0 and h > 0:
        return (float(l) * float(w) * float(h)) / 1_000_000.0
    return None


def to_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def is_pipe_item(name: str, material: str = "") -> bool:
    n = (name or "").upper()
    if "PIPE" in n or "PEX-B" == n.strip() or n.strip() == "PEX-B":
        return True
    if "INSULATION" in n:
        return True
    if material.upper() in {"PEX"} and "ELBOW" not in n and "REDUCER" not in n and "TEE" not in n:
        # PEX-B size rows without fitting words
        if "PEX" in n and "FITTING" not in n:
            return "ELBOW" not in n and "REDUCER" not in n
    return False


def classify_hts(material: str, name: str) -> str:
    mat = (material or "").upper()
    n = (name or "").upper()
    if "INSULATION" in n or mat == "INSULATION":
        return "INSULATION"
    if mat == "COPPER" or "COPPER" in n:
        return "COPPER_PIPE" if is_pipe_item(n, mat) else "COPPER_FITTING"
    if mat == "PEX" or n.startswith("PEX"):
        return "PEX_PIPE" if is_pipe_item(n, mat) else "PEX_FITTING"
    if mat == "CPVC" or "CPVC" in n:
        return "CPVC_PIPE_RIGID" if is_pipe_item(n, mat) else "CPVC_FITTING"
    # PVC default
    if is_pipe_item(n, mat):
        return "PVC_PIPE_RIGID"
    # SCH40 elbows etc from projects may be pressure fittings
    if "SOC" in n or "SCH" in n or "ELL (" in n:
        return "PVC_PRESSURE_FITTING"
    return "PVC_DWV_FITTING"


def hts_lookup(category: str):
    for cat, hts, mfn, notes in HTS_TABLE:
        if cat == category:
            total = mfn + ADD_SEC301 + ADD_FLIP
            return {
                "hts": hts,
                "mfn": mfn,
                "sec301": ADD_SEC301,
                "flip": ADD_FLIP,
                "total": total,
                "notes": notes,
            }
    return {
        "hts": "",
        "mfn": None,
        "sec301": ADD_SEC301,
        "flip": ADD_FLIP,
        "total": None,
        "notes": "Unclassified — verify",
    }


# Website description -> factory name aliases
NAME_ALIASES = {
    "1/16 BEND (H X H)": "1/16 BEND (H X H)",
    "1/16 BEND (H x H)": "1/16 BEND (H X H)",
    "1/16 BEND, STREET (H X S)": "1/16 BEND, STREET (H X S)",
    "1/16 BEND, STREET (H x S)": "1/16 BEND, STREET (H X S)",
    "1/4 BEND (H X H)": "1/4 BEND (H X H)",
    "1/4 BEND (H x H)": "1/4 BEND (H X H)",
    "1/4 BEND, STREET (S X H)": "1/4 BEND, STREET (S X H)",
    "1/4 BEND, STREET (S x H)": "1/4 BEND, STREET (S X H)",
    "1/8 BEND (H X H)": "1/8 BEND (H X H)",
    "1/8 BEND (H x H)": "1/8 BEND (H X H)",
    "1/8 BEND, STREET (H X S)": "1/8 BEND, STREET (H X S)",
    "1/8 BEND, STREET (H x S)": "1/8 BEND, STREET (H X S)",
    "ELBOW 90": "ELBOW 90",
    "COPPER PIPE K": "COPPER PIPE K",
    "COPPER PIPE L": "COPPER PIPE L",
}


def canonical_name(s: str) -> str:
    n = norm_name(s)
    n = n.replace("×", "X")
    n = re.sub(r"\(H\s*X\s*H\)", "(H X H)", n)
    n = re.sub(r"\(H\s*X\s*S\)", "(H X S)", n)
    n = re.sub(r"\(S\s*X\s*H\)", "(S X H)", n)
    n = re.sub(r"\(ALL\s*HUB\)", "(ALL HUB)", n)
    n = re.sub(r"REDUCING SANITARY TEE\s*\(ALL HUB\)", "REDUCING SANITARY TEE (ALL HUB)", n)
    n = re.sub(r"REDUCING WYE\s*\(ALL HUB\)", "REDUCING WYE (ALL HUB)", n)
    # ASTM spacing variants
    n = re.sub(r"ASTM\s+D\s*1785", "ASTM D1785", n)
    n = re.sub(r"ASTM\s+F\s*891", "ASTM F891", n)
    n = re.sub(r"ASTM\s+CPVC", "ASTM CPVC", n)
    n = re.sub(r"\s+", " ", n).strip()
    mapping = {
        "PEX-B": "PEX-B PIPE",
        "ASTM D1785 SCH40 PVC": "ASTM D1785 SCH40 PVC",
        "ASTM F891 PVC FOAM CORE DWV PIPE": "ASTM F891 PVC FOAM CORE DWV PIPE",
        "ASTM CPVC SCH80 PIPE": "ASTM CPVC SCH80 PIPE",
    }
    return mapping.get(n, n)


def load_website():
    rows = []
    with PRODUCTS_CSV.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append(r)
    by_key = {}
    by_tommur_size = defaultdict(list)
    for r in rows:
        name = canonical_name(r.get("Description") or "")
        size = norm_size(r.get("Size"))
        mat = (r.get("Material") or "").strip()
        key = (mat.upper(), name, size)
        by_key[key] = r
        tcode = norm_text(r.get("Tommur-Code") or "")
        # Tommur-Code sometimes includes description for copper
        t_short = tcode.split(" - ")[0].strip() if " - " in tcode else tcode
        by_tommur_size[(t_short.upper(), size)].append(r)
        by_tommur_size[(tcode.upper(), size)].append(r)
    return rows, by_key, by_tommur_size


def load_lesso():
    wb = openpyxl.load_workbook(UPLOADS / "Lesso_DWV_List_f6b7.xlsx", data_only=True)
    ws = wb["Sheet4"]
    by_name_size = {}
    by_code = {}
    for r in range(3, ws.max_row + 1):
        code = norm_text(ws.cell(r, 3).value)
        size = norm_size(ws.cell(r, 4).value)
        pname = canonical_name(ws.cell(r, 5).value or ws.cell(r, 2).value)
        wt = to_float(ws.cell(r, 6).value)
        L, W, H = parse_carton_cm(ws.cell(r, 7).value)
        pcs = to_float(ws.cell(r, 8).value)
        rec = {
            "lesso_code": code,
            "name": pname,
            "size": size,
            "wt_g": wt,
            "L": L,
            "W": W,
            "H": H,
            "pcs_ctn": pcs,
            "cbm_ctn": cbm_from_lwh(L, W, H),
        }
        by_name_size[(pname, size)] = rec
        if code:
            by_code[code] = rec
    return by_name_size, by_code


def plausible_ddp(ddp, fob=None):
    """Reject misplaced sell prices (e.g. Margins pipe rows with DDP=10/20/30)."""
    if ddp is None:
        return False
    if fob is not None and fob > 0:
        ratio = ddp / fob
        # Real DDP in these sheets is ~1.05–1.25x yellow FOB for fittings
        if ratio > 4.0:
            return False
        if ratio < 0.5:
            return False
    elif ddp >= 5:
        # Without FOB context, large round numbers are usually not unit DDP
        return False
    return True


def load_all3_and_margins():
    """Yellow FOB + DDP + dims from All 3 Projects and Margins."""
    by_key = {}

    def ingest(path, yellow_col, list_col=None, ddp_col=None, has_dims=False):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in range(2, ws.max_row + 1):
            mat = norm_text(ws.cell(r, 3).value)
            tcode = norm_text(ws.cell(r, 4).value)
            name = canonical_name(ws.cell(r, 5).value)
            size = norm_size(ws.cell(r, 6).value)
            if not name and not tcode:
                continue
            pcs = to_float(ws.cell(r, 7).value)
            rec = by_key.get((tcode, name, size), {})
            rec.update(
                {
                    "material": mat or rec.get("material"),
                    "tommur_code": tcode or rec.get("tommur_code"),
                    "name": name or rec.get("name"),
                    "size": size or rec.get("size"),
                    "pcs_ctn": pcs or rec.get("pcs_ctn"),
                }
            )
            y = to_float(ws.cell(r, yellow_col).value)
            if y is not None:
                rec["fob_yellow"] = y
            if list_col:
                lv = to_float(ws.cell(r, list_col).value)
                if lv is not None:
                    rec["fob_list"] = lv
            if ddp_col:
                dv = to_float(ws.cell(r, ddp_col).value)
                fob_for_check = y if y is not None else rec.get("fob_yellow")
                if plausible_ddp(dv, fob_for_check):
                    rec["ddp"] = dv
            if has_dims:
                L = to_float(ws.cell(r, 8).value)
                W = to_float(ws.cell(r, 9).value)
                H = to_float(ws.cell(r, 10).value)
                wt = to_float(ws.cell(r, 11).value)
                if L:
                    rec["L"] = L
                if W:
                    rec["W"] = W
                if H:
                    rec["H"] = H
                if wt:
                    rec["wt_g"] = wt
            by_key[(tcode, name, size)] = rec

    ingest(
        UPLOADS / "All_3_Projects_8-27_e6e7.xlsx",
        yellow_col=18,
        list_col=17,
        ddp_col=20,
        has_dims=True,
    )
    ingest(
        UPLOADS / "Margins_8-24-26_284d.xlsx",
        yellow_col=11,
        list_col=None,
        ddp_col=13,
        has_dims=False,
    )
    return by_key


def load_davenport():
    """Parse Davenport PI; forward-fill code/item for size continuation rows."""
    wb = openpyxl.load_workbook(UPLOADS / "Davenport_8-27_687f.xlsx", data_only=True)
    ws = wb["Sheet0"]
    rows = []
    cur_code = None
    cur_item = None
    for r in range(13, ws.max_row + 1):
        code = ws.cell(r, 1).value
        item = ws.cell(r, 3).value
        size = ws.cell(r, 4).value
        if code and str(code).strip().lower() in {"total", "the total price"}:
            break
        if isinstance(code, str) and code.strip().startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
            break
        if code:
            cur_code = norm_text(code)
        if item:
            cur_item = canonical_name(item)
        if size is None and to_float(ws.cell(r, 10).value) is None:
            continue
        # infer item from picture/remarks if needed — picture col may have text in some exports
        pic = ws.cell(r, 2).value
        if pic and isinstance(pic, str) and not cur_item:
            cur_item = canonical_name(pic)
        fob = to_float(ws.cell(r, 10).value)
        rec = {
            "tommur_code": cur_code,
            "name": cur_item,
            "size": norm_size(size),
            "pcs_ctn": to_float(ws.cell(r, 5).value),
            "L": to_float(ws.cell(r, 6).value),
            "W": to_float(ws.cell(r, 7).value),
            "H": to_float(ws.cell(r, 8).value),
            "wt_g": to_float(ws.cell(r, 9).value),
            "fob_dav": fob,
            "cbm_line": to_float(ws.cell(r, 14).value),
        }
        if rec["size"] or rec["fob_dav"] is not None:
            rows.append(rec)
    return rows


def load_all_current():
    wb = openpyxl.load_workbook(UPLOADS / "All_Current_Supplies_9451.xlsx", data_only=True)
    out = []

    # PVC: ITEM, SIZE, PCS/CTN, FOB
    ws = wb["PVC"]
    for r in range(2, ws.max_row + 1):
        name = canonical_name(ws.cell(r, 1).value)
        if not name:
            continue
        out.append(
            {
                "material": "PVC",
                "tommur_code": "",
                "name": name,
                "size": norm_size(ws.cell(r, 2).value),
                "pcs_ctn": to_float(ws.cell(r, 3).value),
                "fob_acs": to_float(ws.cell(r, 4).value),
                "source_sheet": "All_Current_Supplies!PVC",
            }
        )

    # CPVC: Code, ITEM, SIZE, PCS/CTN, FOB
    ws = wb["CPVC"]
    for r in range(2, ws.max_row + 1):
        tcode = norm_text(ws.cell(r, 1).value).replace("\n", " ")
        tcode = re.sub(r"\s+", " ", tcode)
        name = canonical_name(ws.cell(r, 2).value)
        if not name and not tcode:
            continue
        out.append(
            {
                "material": "CPVC",
                "tommur_code": tcode,
                "name": name,
                "size": norm_size(ws.cell(r, 3).value),
                "pcs_ctn": to_float(ws.cell(r, 4).value),
                "fob_acs": to_float(ws.cell(r, 5).value),
                "source_sheet": "All_Current_Supplies!CPVC",
            }
        )

    # Copper: Code, Description, Size, Pack, Price (= website sell, NOT FOB)
    ws = wb["Copper"]
    for r in range(2, ws.max_row + 1):
        apbs = norm_text(ws.cell(r, 1).value)
        name = canonical_name(ws.cell(r, 2).value)
        if not name:
            continue
        out.append(
            {
                "material": "Copper",
                "tommur_code": "",
                "apbs_code_hint": apbs,
                "name": name,
                "size": norm_size(ws.cell(r, 3).value),
                "pcs_ctn": to_float(ws.cell(r, 4).value),
                "sell_acs": to_float(ws.cell(r, 5).value),  # selling price copy
                "source_sheet": "All_Current_Supplies!Copper",
            }
        )

    # PEX: Code, ITEM, SIZE, PCS/CTN, FOB
    ws = wb["PEX"]
    for r in range(2, ws.max_row + 1):
        tcode = norm_text(ws.cell(r, 1).value)
        name = canonical_name(ws.cell(r, 2).value)
        if not name:
            continue
        out.append(
            {
                "material": "PEX",
                "tommur_code": tcode,
                "name": name if name != "PEX-B" else "PEX-B PIPE",
                "size": norm_size(ws.cell(r, 3).value),
                "pcs_ctn": to_float(ws.cell(r, 4).value),
                "fob_acs": to_float(ws.cell(r, 5).value),
                "source_sheet": "All_Current_Supplies!PEX",
            }
        )

    return out


def merge_all():
    web_rows, web_by_key, web_by_tommur = load_website()
    lesso_by_ns, lesso_by_code = load_lesso()
    proj = load_all3_and_margins()
    dav = load_davenport()
    acs = load_all_current()

    # Index project by multiple keys
    proj_by_tcode_size = defaultdict(list)
    proj_by_name_size = defaultdict(list)
    for (tcode, name, size), rec in proj.items():
        proj_by_tcode_size[(norm_text(tcode).upper(), size)].append(rec)
        proj_by_name_size[(canonical_name(name), size)].append(rec)

    dav_by_tcode_size = defaultdict(list)
    dav_by_name_size = defaultdict(list)
    for rec in dav:
        dav_by_tcode_size[(norm_text(rec.get("tommur_code")).upper(), rec.get("size"))].append(rec)
        if rec.get("name"):
            dav_by_name_size[(canonical_name(rec["name"]), rec.get("size"))].append(rec)

    # Seed universe from All Current Supplies + any project-only + website-only Tommur items
    universe = {}

    def ukey(material, name, size, tcode=""):
        return (material.upper(), canonical_name(name), norm_size(size), norm_text(tcode).upper())

    def ensure(material, name, size, tcode=""):
        k = ukey(material, name, size, tcode)
        if k not in universe:
            universe[k] = {
                "material": material,
                "name": canonical_name(name),
                "size": norm_size(size),
                "tommur_code": norm_text(tcode),
            }
        return universe[k]

    for row in acs:
        rec = ensure(row["material"], row["name"], row["size"], row.get("tommur_code") or "")
        for fld in ("pcs_ctn", "fob_acs", "sell_acs", "apbs_code_hint", "source_sheet"):
            if row.get(fld) not in (None, ""):
                rec[fld] = row[fld]

    # Add project rows not already present
    for (tcode, name, size), prec in proj.items():
        mat = prec.get("material") or "UNKNOWN"
        # try match existing without tcode first for PVC ACS
        found = None
        for k, rec in universe.items():
            if k[0] == mat.upper() and k[1] == canonical_name(name) and k[2] == norm_size(size):
                found = rec
                break
        if found is None:
            found = ensure(mat, name, size, tcode)
        if tcode and not found.get("tommur_code"):
            found["tommur_code"] = tcode
        for fld in ("fob_yellow", "fob_list", "ddp", "L", "W", "H", "wt_g", "pcs_ctn"):
            if prec.get(fld) is not None:
                found[fld] = prec[fld]

    # Website rows (adds insulation etc. and fills codes)
    for wr in web_rows:
        mat = wr.get("Material") or ""
        name = canonical_name(wr.get("Description") or "")
        size = norm_size(wr.get("Size"))
        tcode_raw = norm_text(wr.get("Tommur-Code") or "")
        tcode = tcode_raw.split(" - ")[0].strip() if " - " in tcode_raw else tcode_raw
        # copper website tommur includes type in code field like "Copper Type K Soft - ELBOW 90"
        found = None
        for k, rec in universe.items():
            if k[0] == mat.upper() and k[1] == name and k[2] == size:
                found = rec
                break
        if found is None:
            # try by tommur+size
            for k, rec in universe.items():
                if norm_size(k[2]) == size and (
                    norm_text(rec.get("tommur_code")).upper() == tcode.upper()
                    or norm_text(rec.get("apbs_code_hint")) == wr.get("Code")
                ):
                    # name compatibility for copper
                    if mat.upper() == "COPPER" or k[0] == "COPPER":
                        found = rec
                        break
        if found is None:
            found = ensure(mat, name, size, tcode)
        found["apbs_code"] = wr.get("Code")
        found["lesso_code_web"] = wr.get("Lesso-Code") or ""
        found["sell_web"] = to_float(wr.get("Price"))
        found["pack_web"] = to_float(wr.get("Pack"))
        if tcode and not found.get("tommur_code"):
            found["tommur_code"] = tcode
        if tcode_raw:
            found["tommur_code_web"] = tcode_raw

    # Enrich from Lesso by name+size
    for rec in universe.values():
        ns = (rec["name"], rec["size"])
        # try direct
        Lrec = lesso_by_ns.get(ns)
        # try alternate name forms
        if not Lrec:
            alt = rec["name"].replace(" (ALL HUB)", "(ALL HUB)").replace("(ALL HUB)", " (ALL HUB)")
            alt = re.sub(r"\s+", " ", alt).strip()
            Lrec = lesso_by_ns.get((canonical_name(alt), rec["size"]))
        if not Lrec and rec.get("lesso_code_web"):
            Lrec = lesso_by_code.get(rec["lesso_code_web"])
        if Lrec:
            rec["lesso_code"] = Lrec.get("lesso_code") or rec.get("lesso_code_web") or ""
            for fld in ("L", "W", "H", "pcs_ctn", "wt_g"):
                if rec.get(fld) in (None, "") and Lrec.get(fld) is not None:
                    rec[fld] = Lrec[fld]
        elif rec.get("lesso_code_web"):
            rec["lesso_code"] = rec["lesso_code_web"]

    # Enrich from project by tommur+size / name+size if missing FOB/DDP/dims
    for rec in universe.values():
        cands = []
        tc = norm_text(rec.get("tommur_code")).upper()
        if tc:
            cands.extend(proj_by_tcode_size.get((tc, rec["size"]), []))
        cands.extend(proj_by_name_size.get((rec["name"], rec["size"]), []))
        for prec in cands:
            for fld in ("fob_yellow", "fob_list", "ddp", "L", "W", "H", "wt_g", "pcs_ctn"):
                if rec.get(fld) in (None, "") and prec.get(fld) is not None:
                    rec[fld] = prec[fld]
            if not rec.get("tommur_code") and prec.get("tommur_code"):
                rec["tommur_code"] = prec["tommur_code"]
            if not rec.get("material") and prec.get("material"):
                rec["material"] = prec["material"]

        dcands = []
        if tc:
            dcands.extend(dav_by_tcode_size.get((tc, rec["size"]), []))
        dcands.extend(dav_by_name_size.get((rec["name"], rec["size"]), []))
        for drec in dcands:
            if rec.get("fob_dav") is None and drec.get("fob_dav") is not None:
                rec["fob_dav"] = drec["fob_dav"]
            for fld in ("L", "W", "H", "wt_g", "pcs_ctn"):
                if rec.get(fld) in (None, "") and drec.get(fld) is not None:
                    rec[fld] = drec[fld]
            if not rec.get("tommur_code") and drec.get("tommur_code"):
                rec["tommur_code"] = drec["tommur_code"]

    # Website tommur mapping for PVC codes like D035 when ACS lacked them
    for rec in universe.values():
        if rec.get("tommur_code"):
            continue
        matches = web_by_tommur.get(("", ""))  # noop
        for wr in web_rows:
            if (wr.get("Material") or "").upper() != (rec.get("material") or "").upper():
                continue
            if canonical_name(wr.get("Description")) != rec["name"]:
                continue
            if norm_size(wr.get("Size")) != rec["size"]:
                continue
            tcode_raw = norm_text(wr.get("Tommur-Code") or "")
            tcode = tcode_raw.split(" - ")[0].strip() if " - " in tcode_raw else tcode_raw
            rec["tommur_code"] = tcode
            rec["apbs_code"] = wr.get("Code")
            rec["lesso_code"] = rec.get("lesso_code") or wr.get("Lesso-Code") or ""
            rec["sell_web"] = to_float(wr.get("Price"))
            break

    # Final dedupe: merge rows that share material+name+size and compatible tommur codes
    merged = {}
    for rec in universe.values():
        mat = (rec.get("material") or "").upper()
        name = canonical_name(rec.get("name") or "")
        size = norm_size(rec.get("size") or "")
        tc = norm_text(rec.get("tommur_code") or "").upper()
        # Prefer matching an existing row with same mat/name/size
        match_k = None
        for k, existing in merged.items():
            if k[0] != mat or k[1] != name or k[2] != size:
                continue
            etc = norm_text(existing.get("tommur_code") or "").upper()
            if not tc or not etc or tc == etc or tc in etc or etc in tc:
                match_k = k
                break
            # Copper: match on APBS code hint
            if mat == "COPPER":
                a1 = norm_text(rec.get("apbs_code") or rec.get("apbs_code_hint") or "")
                a2 = norm_text(existing.get("apbs_code") or existing.get("apbs_code_hint") or "")
                if a1 and a2 and a1 == a2:
                    match_k = k
                    break
        if match_k is None:
            merged[(mat, name, size, tc)] = rec
        else:
            dest = merged[match_k]
            for fld, val in rec.items():
                if val in (None, ""):
                    continue
                if dest.get(fld) in (None, ""):
                    dest[fld] = val
                elif fld == "tommur_code" and len(str(val)) > len(str(dest.get(fld) or "")):
                    # keep more specific code when compatible
                    if not dest.get(fld):
                        dest[fld] = val

    return list(merged.values())


def finalize_row(rec):
    material = rec.get("material") or ""
    name = rec.get("name") or ""
    pipe = is_pipe_item(name, material)
    sell_unit = "per ft" if pipe and "INSULATION" not in name.upper() else "per pc"
    # insulation often per pc stick/tube — keep per pc; foam pipe on site is length pricing sometimes
    if "PIPE" in name.upper() and material.upper() in {"PVC", "CPVC", "PEX", "COPPER"}:
        sell_unit = "per ft"
    if "INSULATION" in name.upper():
        sell_unit = "per ft"

    L, W, H = rec.get("L"), rec.get("W"), rec.get("H")
    pcs = rec.get("pcs_ctn") or rec.get("pack_web")
    cbm_ctn = cbm_from_lwh(L, W, H)
    cbm_pc = (cbm_ctn / pcs) if cbm_ctn and pcs and pcs > 0 else None

    # Prefer yellow FOB; else list; else davenport; else ACS (may be DDP-like for PVC)
    fob = None
    fob_source = ""
    if rec.get("fob_yellow") is not None:
        fob = rec["fob_yellow"]
        fob_source = "Yellow FOB (All3/Margins = list x 0.95)"
    elif rec.get("fob_list") is not None:
        fob = rec["fob_list"]
        fob_source = "FOB list (All3 col Q / pre-discount)"
    elif rec.get("fob_dav") is not None:
        fob = rec["fob_dav"]
        fob_source = "Davenport PI FOB"
    elif rec.get("fob_acs") is not None and material.upper() != "COPPER":
        fob = rec["fob_acs"]
        fob_source = "All Current Supplies FOB (verify — some PVC rows track old DDP)"

    ddp = rec.get("ddp")
    ddp_note = ""
    # Infer DDP from All Current Supplies when it sits above yellow FOB in the
    # historical ~1.05–1.25x band (ACS PVC "FOB" often stored old DDP).
    if (
        ddp is None
        and rec.get("fob_acs") is not None
        and rec.get("fob_yellow") is not None
        and material.upper() == "PVC"
    ):
        ratio = rec["fob_acs"] / rec["fob_yellow"] if rec["fob_yellow"] else None
        if ratio and 1.02 <= ratio <= 1.35:
            ddp = rec["fob_acs"]
            ddp_note = "DDP inferred from All Current Supplies (matches prior DDP vs yellow FOB pattern)"
    elif ddp is None and rec.get("fob_acs") is not None and rec.get("fob_yellow") is not None:
        ddp_note = f"ACS value {rec['fob_acs']} present (not copied to DDP; ratio vs yellow outside 1.02–1.35)"

    sell = rec.get("sell_web")
    sell_source = "Website" if sell is not None else ""
    if sell is None and rec.get("sell_acs") is not None and material.upper() == "COPPER":
        sell = rec["sell_acs"]
        sell_source = "All Current Supplies Copper Price (= site sell)"

    cat = classify_hts(material, name)
    h = hts_lookup(cat)

    freight_40 = (FREIGHT_PER_CONTAINER / CBM_40FT) * cbm_pc if cbm_pc else None
    freight_45 = (FREIGHT_PER_CONTAINER / CBM_45HQ) * cbm_pc if cbm_pc else None

    duty_amt = (fob * h["total"]) if (fob is not None and h["total"] is not None) else None
    landed_40 = (fob + duty_amt + freight_40) if (fob is not None and duty_amt is not None and freight_40 is not None) else None
    landed_45 = (fob + duty_amt + freight_45) if (fob is not None and duty_amt is not None and freight_45 is not None) else None
    # if dims missing but have fob+duty, still show partial
    if landed_40 is None and fob is not None and duty_amt is not None and freight_40 is None:
        landed_40 = None  # leave blank — incomplete
    if fob is not None and duty_amt is not None:
        landed_ex_freight = fob + duty_amt
    else:
        landed_ex_freight = None

    margin_40 = (sell - landed_40) if (sell is not None and landed_40 is not None) else None
    margin_45 = (sell - landed_45) if (sell is not None and landed_45 is not None) else None
    margin_pct_40 = (margin_40 / sell) if (margin_40 is not None and sell) else None
    margin_pct_45 = (margin_45 / sell) if (margin_45 is not None and sell) else None

    fob_vs_ddp = None
    fob_lower = ""
    if fob is not None and ddp is not None:
        fob_vs_ddp = ddp - fob
        fob_lower = "YES" if fob < ddp else ("SAME" if abs(fob - ddp) < 1e-9 else "NO — FOB >= DDP")

    return {
        "APBS_Item_Code": rec.get("apbs_code") or rec.get("apbs_code_hint") or "",
        "Tommur_Code": rec.get("tommur_code") or "",
        "Lesso_Code": rec.get("lesso_code") or rec.get("lesso_code_web") or "",
        "Material": material,
        "Description": name,
        "Size": rec.get("size") or "",
        "Color": "",
        "Sell_Unit": sell_unit,
        "Pcs_per_Carton": pcs,
        "Carton_L_cm": L,
        "Carton_W_cm": W,
        "Carton_H_cm": H,
        "CBM_per_Carton": cbm_ctn,
        "CBM_per_Pc": cbm_pc,
        "FOB_USD": fob,
        "FOB_Source": fob_source,
        "FOB_Yellow": rec.get("fob_yellow"),
        "FOB_List": rec.get("fob_list"),
        "FOB_Davenport": rec.get("fob_dav"),
        "FOB_AllCurrentSupplies": rec.get("fob_acs") if material.upper() != "COPPER" else None,
        "DDP_Current": ddp,
        "FOB_vs_DDP_Savings": fob_vs_ddp,
        "FOB_Lower_Than_DDP": fob_lower,
        "Selling_Price": sell,
        "Sell_Price_Source": sell_source,
        "HTS_Code": h["hts"],
        "HTS_Category": cat,
        "Duty_MFN_Pct": h["mfn"],
        "Tariff_Sec301_Pct": h["sec301"],
        "Tariff_FLIP301_Pct": h["flip"],
        "Duty_Tariff_Total_Pct": h["total"],
        "Est_Duty_Tariff_USD": duty_amt,
        "Freight_per_Pc_40ft": freight_40,
        "Freight_per_Pc_45HQ": freight_45,
        "Est_Landed_per_Pc_40ft": landed_40,
        "Est_Landed_per_Pc_45HQ": landed_45,
        "Landed_ex_Freight": landed_ex_freight,
        "Margin_USD_40ft": margin_40,
        "Margin_USD_45HQ": margin_45,
        "Margin_Pct_40ft": margin_pct_40,
        "Margin_Pct_45HQ": margin_pct_45,
        "Net_Wt_g_per_Pc": rec.get("wt_g"),
        "Notes": ddp_note,
        "On_Website": "YES" if rec.get("apbs_code") or (rec.get("sell_web") is not None) else "NO",
    }


HEADERS = [
    "APBS_Item_Code",
    "Tommur_Code",
    "Lesso_Code",
    "Material",
    "Description",
    "Size",
    "Color",
    "Sell_Unit",
    "Pcs_per_Carton",
    "Carton_L_cm",
    "Carton_W_cm",
    "Carton_H_cm",
    "CBM_per_Carton",
    "CBM_per_Pc",
    "FOB_USD",
    "FOB_Source",
    "FOB_Yellow",
    "FOB_List",
    "FOB_Davenport",
    "FOB_AllCurrentSupplies",
    "DDP_Current",
    "FOB_vs_DDP_Savings",
    "FOB_Lower_Than_DDP",
    "Selling_Price",
    "Sell_Price_Source",
    "HTS_Code",
    "HTS_Category",
    "Duty_MFN_Pct",
    "Tariff_Sec301_Pct",
    "Tariff_FLIP301_Pct",
    "Duty_Tariff_Total_Pct",
    "Est_Duty_Tariff_USD",
    "Freight_per_Pc_40ft",
    "Freight_per_Pc_45HQ",
    "Est_Landed_per_Pc_40ft",
    "Est_Landed_per_Pc_45HQ",
    "Landed_ex_Freight",
    "Margin_USD_40ft",
    "Margin_USD_45HQ",
    "Margin_Pct_40ft",
    "Margin_Pct_45HQ",
    "Net_Wt_g_per_Pc",
    "On_Website",
    "Notes",
]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True, size=10)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin
    ws.row_dimensions[row].height = 36
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width=28):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = 0
        for row in range(1, min(ws.max_row, 80) + 1):
            v = ws.cell(row, col).value
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[letter].width = min(max(10, maxlen + 2), max_width)


def write_workbook(rows):
    wb = openpyxl.Workbook()

    # --- Assumptions ---
    ws = wb.active
    ws.title = "Assumptions_Notes"
    notes = [
        ["Tommur Cost & Margin Tracker", ""],
        ["Generated", date.today().isoformat()],
        ["", ""],
        ["FREIGHT ASSUMPTIONS", ""],
        ["Freight per container (USD)", FREIGHT_PER_CONTAINER],
        ["40' container CBM (internal)", CBM_40FT],
        ["45'HQ container CBM (internal)", CBM_45HQ],
        ["Freight $/CBM 40'", round(FREIGHT_PER_CONTAINER / CBM_40FT, 4)],
        ["Freight $/CBM 45'HQ", round(FREIGHT_PER_CONTAINER / CBM_45HQ, 4)],
        ["Freight per pc formula", "(7000 / container_CBM) * CBM_per_pc"],
        ["CBM_per_pc formula", "(L_cm * W_cm * H_cm / 1,000,000) / pcs_per_carton"],
        ["", ""],
        ["FOB / DDP RULES", ""],
        ["Primary FOB (FOB_USD)", "Yellow FOB from All 3 Projects / Margins (= listed FOB x 0.95)"],
        ["Fallback FOB order", "FOB list → Davenport PI → All Current Supplies FOB"],
        ["DDP_Current", "From project sheets 'Current DDP' (first shipments). Shown for FOB vs DDP compare."],
        ["ACS PVC note", "All Current Supplies PVC 'FOB' often matches old DDP — surfaced in FOB_AllCurrentSupplies; only used as FOB_USD if no yellow/list/PI."],
        ["", ""],
        ["LANDED COST", ""],
        ["Formula", "FOB + (FOB * total_duty_tariff_pct) + freight_per_pc"],
        ["", ""],
        ["SELLING PRICE", ""],
        ["Source priority", "Website products.csv; Copper ACS price used when it matches site sell"],
        ["MISSING FILE", "Complete Fittings Price Lists.xlsx was referenced from local OneDrive but not uploaded — upload it to refresh fitting sell prices"],
        ["Blanks", "Where no sell price exists, Selling_Price left blank for you to set"],
        ["", ""],
        ["HTS / DUTY / TARIFF (estimate — verify with customs broker before entry)", ""],
        ["MFN", "Ordinary HTS Column 1 general rate"],
        ["Section 301", f"{ADD_SEC301*100:.1f}% additional on China-origin for these headings (Lists 1-3 typical)"],
        ["FLIP 301 / forced-labor add-on", f"{ADD_FLIP*100:.1f}% additional on China products of China effective ~2026-07-24 (USTR FLIP 301 action) — confirm applicability per entry"],
        ["Total stack used here", "MFN + 25% Sec 301 + 12.5% FLIP"],
        ["NOT included", "MPF (0.3464%), HMF (0.125%), brokerage, trucking, Section 232 copper (if any) — add if needed"],
        ["Color", "Left blank (not in source files)"],
        ["", ""],
        ["DATA SOURCES", ""],
        ["All_3_Projects_8-27.xlsx", "Yellow FOB, list FOB, DDP, carton dims, Tommur codes"],
        ["Margins_8-24-26.xlsx", "Yellow FOB confirmation"],
        ["Davenport_8-27.xlsx", "Tommur PI dims / alternate FOB"],
        ["All_Current_Supplies.xlsx", "Full Tommur-offerable catalog (PVC/CPVC/Copper/PEX)"],
        ["Lesso_DWV_List.xlsx", "Lesso codes + carton LxWxH + pcs/ctn for DWV fittings"],
        ["Website assets/products.csv", "APBS item codes, Lesso codes, selling prices"],
    ]
    for r in notes:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 100

    # --- HTS reference ---
    ws = wb.create_sheet("HTS_Tariff_Reference")
    ws.append(
        [
            "HTS_Category",
            "HTS_Code",
            "MFN_Pct",
            "Sec301_Pct",
            "FLIP301_Pct",
            "Total_Pct",
            "Description_Notes",
        ]
    )
    for cat, hts, mfn, desc in HTS_TABLE:
        total = mfn + ADD_SEC301 + ADD_FLIP
        ws.append([cat, hts, mfn, ADD_SEC301, ADD_FLIP, total, desc])
    style_header(ws)
    for r in range(2, ws.max_row + 1):
        for c in (3, 4, 5, 6):
            ws.cell(r, c).number_format = "0.0%"
    autosize(ws, 55)

    # --- Master ---
    ws = wb.create_sheet("Cost_Margin_Master", 0)
    # move assumptions after? Keep master first
    # Actually created Assumptions first; move master to front:
    wb.move_sheet(ws, offset=-len(wb.sheetnames) + 1)

    ws.append(HEADERS)
    # Sort: Material, Description, Size
    rows_sorted = sorted(
        rows,
        key=lambda r: (
            r.get("Material") or "",
            r.get("Description") or "",
            r.get("Size") or "",
            r.get("Tommur_Code") or "",
        ),
    )
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    money_cols = {
        "FOB_USD",
        "FOB_Yellow",
        "FOB_List",
        "FOB_Davenport",
        "FOB_AllCurrentSupplies",
        "DDP_Current",
        "FOB_vs_DDP_Savings",
        "Selling_Price",
        "Est_Duty_Tariff_USD",
        "Freight_per_Pc_40ft",
        "Freight_per_Pc_45HQ",
        "Est_Landed_per_Pc_40ft",
        "Est_Landed_per_Pc_45HQ",
        "Landed_ex_Freight",
        "Margin_USD_40ft",
        "Margin_USD_45HQ",
    }
    pct_cols = {
        "Duty_MFN_Pct",
        "Tariff_Sec301_Pct",
        "Tariff_FLIP301_Pct",
        "Duty_Tariff_Total_Pct",
        "Margin_Pct_40ft",
        "Margin_Pct_45HQ",
    }
    for rec in rows_sorted:
        ws.append([rec.get(h) for h in HEADERS])
    style_header(ws)

    header_idx = {h: i + 1 for i, h in enumerate(HEADERS)}
    for r in range(2, ws.max_row + 1):
        for name in money_cols:
            ws.cell(r, header_idx[name]).number_format = "0.0000"
        for name in pct_cols:
            ws.cell(r, header_idx[name]).number_format = "0.0%"
        for name in ("CBM_per_Carton", "CBM_per_Pc"):
            ws.cell(r, header_idx[name]).number_format = "0.000000"
        # highlight yellow FOB used
        if ws.cell(r, header_idx["FOB_Source"]).value and "Yellow" in str(
            ws.cell(r, header_idx["FOB_Source"]).value
        ):
            ws.cell(r, header_idx["FOB_USD"]).fill = yellow_fill
        # flag FOB not lower than DDP
        flag = ws.cell(r, header_idx["FOB_Lower_Than_DDP"]).value
        if flag and str(flag).startswith("NO"):
            ws.cell(r, header_idx["FOB_Lower_Than_DDP"]).fill = PatternFill(
                "solid", fgColor="F4CCCC"
            )
        elif flag == "YES":
            ws.cell(r, header_idx["FOB_Lower_Than_DDP"]).fill = PatternFill(
                "solid", fgColor="D9EAD3"
            )

    autosize(ws, 22)
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["P"].width = 36

    # --- FOB vs DDP compare subset ---
    ws = wb.create_sheet("FOB_vs_DDP_Compare")
    cmp_headers = [
        "APBS_Item_Code",
        "Tommur_Code",
        "Material",
        "Description",
        "Size",
        "FOB_USD",
        "DDP_Current",
        "FOB_vs_DDP_Savings",
        "FOB_Lower_Than_DDP",
        "FOB_Source",
        "Selling_Price",
    ]
    ws.append(cmp_headers)
    for rec in rows_sorted:
        if rec.get("FOB_USD") is not None and rec.get("DDP_Current") is not None:
            ws.append([rec.get(h) for h in cmp_headers])
    style_header(ws)
    for r in range(2, ws.max_row + 1):
        for c in range(6, 9):
            ws.cell(r, c).number_format = "0.0000"
        flag = ws.cell(r, 9).value
        if flag and str(flag).startswith("NO"):
            ws.cell(r, 9).fill = PatternFill("solid", fgColor="F4CCCC")
        elif flag == "YES":
            ws.cell(r, 9).fill = PatternFill("solid", fgColor="D9EAD3")
    autosize(ws, 36)

    # --- Data gaps ---
    ws = wb.create_sheet("Data_Gaps")
    ws.append(
        [
            "APBS_Item_Code",
            "Tommur_Code",
            "Material",
            "Description",
            "Size",
            "Missing_FOB",
            "Missing_DDP",
            "Missing_Dims",
            "Missing_Sell_Price",
            "Missing_Lesso",
            "On_Website",
        ]
    )
    for rec in rows_sorted:
        miss_fob = "X" if rec.get("FOB_USD") is None else ""
        miss_ddp = "X" if rec.get("DDP_Current") is None else ""
        miss_dims = "X" if rec.get("CBM_per_Pc") is None else ""
        miss_sell = "X" if rec.get("Selling_Price") is None else ""
        miss_lesso = (
            "X"
            if (rec.get("Material") or "").upper() == "PVC"
            and "PIPE" not in (rec.get("Description") or "").upper()
            and not rec.get("Lesso_Code")
            else ""
        )
        if any([miss_fob, miss_dims, miss_sell]):
            ws.append(
                [
                    rec.get("APBS_Item_Code"),
                    rec.get("Tommur_Code"),
                    rec.get("Material"),
                    rec.get("Description"),
                    rec.get("Size"),
                    miss_fob,
                    miss_ddp,
                    miss_dims,
                    miss_sell,
                    miss_lesso,
                    rec.get("On_Website"),
                ]
            )
    style_header(ws)
    autosize(ws, 40)

    # --- Summary counts ---
    ws = wb.create_sheet("Summary")
    n = len(rows_sorted)
    n_fob = sum(1 for r in rows_sorted if r.get("FOB_USD") is not None)
    n_ddp = sum(1 for r in rows_sorted if r.get("DDP_Current") is not None)
    n_both = sum(
        1
        for r in rows_sorted
        if r.get("FOB_USD") is not None and r.get("DDP_Current") is not None
    )
    n_fob_ok = sum(1 for r in rows_sorted if r.get("FOB_Lower_Than_DDP") == "YES")
    n_fob_bad = sum(
        1 for r in rows_sorted if str(r.get("FOB_Lower_Than_DDP") or "").startswith("NO")
    )
    n_sell = sum(1 for r in rows_sorted if r.get("Selling_Price") is not None)
    n_web = sum(1 for r in rows_sorted if r.get("On_Website") == "YES")
    n_dims = sum(1 for r in rows_sorted if r.get("CBM_per_Pc") is not None)
    by_mat = defaultdict(int)
    for r in rows_sorted:
        by_mat[r.get("Material") or ""] += 1
    summary = [
        ["Metric", "Value"],
        ["Total SKUs (Tommur-offerable merged)", n],
        ["On website already", n_web],
        ["Not yet on website", n - n_web],
        ["Have FOB_USD", n_fob],
        ["Have DDP_Current", n_ddp],
        ["Have both FOB + DDP", n_both],
        ["FOB lower than DDP", n_fob_ok],
        ["FOB NOT lower than DDP (review)", n_fob_bad],
        ["Have selling price", n_sell],
        ["Have carton dims / CBM", n_dims],
        ["", ""],
        ["By material", "Count"],
    ]
    for m, c in sorted(by_mat.items()):
        summary.append([m, c])
    for row in summary:
        ws.append(row)
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    autosize(ws, 45)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    return {
        "n": n,
        "n_fob": n_fob,
        "n_ddp": n_ddp,
        "n_both": n_both,
        "n_fob_ok": n_fob_ok,
        "n_fob_bad": n_fob_bad,
        "n_sell": n_sell,
        "n_web": n_web,
        "n_dims": n_dims,
        "by_mat": dict(by_mat),
    }


def main():
    raw = merge_all()
    final = [finalize_row(r) for r in raw]
    # drop empty junk
    final = [r for r in final if r.get("Description") or r.get("Tommur_Code")]
    stats = write_workbook(final)
    print("Wrote", OUT_PATH)
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
