#!/usr/bin/env python3
"""Green Valve PI YMF26060303X (misnamed Zhenpeng PDF) vs Tommur PVC DDP/FOB."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

OUT = Path(__file__).resolve().parent
DUTY = 0.428
# EXW Taizhou → pad ~3% inland to Ningbo to approximate FOB
INLAND = 0.03
FREIGHT_DEFAULT = 0.03

# Green Valve PI 2026/6/3 vs Tommur NJPD landed sheet (FOB / DDP / sell)
# 90deg SxS = Tommur 1/4 BEND HxH. 45deg SxS = Tommur 1/8 BEND HxH.
ROWS = [
    # item, size, gv_exw, gv_qty, tommur_fob, tommur_ddp, sell, freight, match_note
    ("DWV coupling SxS", '1-1/2"', 0.065, 560, 0.1129, 0.12, 0.25, 0.0095, "Tommur COUPLING HxH"),
    ("DWV coupling SxS", '2"', 0.098, 630, 0.1697, 0.17, 0.34, 0.0182, "Tommur COUPLING HxH"),
    ("DWV coupling SxS", '3"', 0.265, 360, 0.5217, 0.55, 1.15, 0.0452, "Tommur COUPLING HxH"),
    ("DWV coupling SxS", '4"', 0.449, 150, 0.8943, 0.95, 2.13, 0.1086, "Tommur COUPLING HxH"),
    ("DWV 90 elbow SxS", '1-1/2"', 0.133, 1150, 0.2307, 0.27, 0.53, 0.0271, "Tommur 1/4 BEND HxH"),
    ("DWV 90 elbow SxS", '2"', 0.203, 1080, 0.3734, 0.43, 0.86, 0.0543, "Tommur 1/4 BEND HxH"),
    ("DWV 90 elbow SxS", '3"', 0.466, 816, 1.1046, 1.27, 2.65, 0.1513, "Tommur 1/4 BEND HxH"),
    ("DWV 90 elbow SxS", '4"', 0.953, 105, 2.1055, 2.22, 4.99, 0.2714, "Tommur 1/4 BEND HxH"),
    ("DWV 45 elbow SxS", '1-1/2"', 0.111, 1000, 0.2131, 0.23, 0.52, 0.0182, "Tommur 1/8 BEND HxH"),
    ("DWV 45 elbow SxS", '2"', 0.166, 1120, 0.3060, 0.35, 0.79, 0.0365, "Tommur 1/8 BEND HxH"),
    ("DWV 45 elbow SxS", '3"', 0.356, 800, 0.9091, 1.05, 2.16, 0.1086, "Tommur 1/8 BEND HxH"),
    ("DWV 45 elbow SxS", '4"', 0.739, 132, 1.6227, 1.87, 4.11, 0.2522, "Tommur 1/8 BEND HxH"),
    ("DWV tee SxSxS", '1-1/2"', 0.183, 387, 0.3459, 0.36, 0.91, 0.0373, "Tommur SANITARY TEE"),
    ("DWV tee SxSxS", '2"', 0.295, 280, 0.4845, 0.52, 1.32, 0.0775, "Tommur SANITARY TEE"),
    ("DWV tee SxSxS", '3"', 0.745, 84, 1.5034, 1.70, 6.99, 0.1809, "Tommur SANITARY TEE"),
    ("DWV tee SxSxS", '4"', 1.324, 20, 2.8505, 3.16, 14.75, 0.4477, "Tommur SANITARY TEE"),
    ("DWV wye SxSxS", '1-1/2"', 0.231, 110, 0.4301, 0.45, 1.72, 0.0448, "Tommur WYE ALL HUB"),
    ("DWV wye SxSxS", '2"', 0.365, 336, 0.5963, 0.63, 3.60, 0.0730, "Tommur WYE ALL HUB"),
    ("DWV wye SxSxS", '3"', 0.939, 162, 1.7967, 2.07, 7.19, 0.2522, "Tommur WYE ALL HUB"),
    ("DWV wye SxSxS", '4"', 1.665, 48, 3.2748, 3.77, 13.10, 0.4729, "Tommur WYE ALL HUB"),
    ("DWV reducing wye 4x4x2", '4x4x2"', 1.026, 26, 1.8964, 2.18, 5.30, 0.2713, "Tommur REDUCING WYE 4x4x2"),
    ("DWV P-trap SxS", '2"', 0.869, 30, 0.9434, 0.99, 2.35, 0.1215, "Tommur P-TRAP HxH"),
    ("DWV plug (socket)", '1-1/2"', 0.159, 280, 0.0837, 0.09, 0.65, 0.0076, "NOT same as Tommur CAP SOC — plug vs cap"),
    ("DWV plug (socket)", '2"', 0.225, 216, 0.1224, 0.13, 1.48, 0.0101, "NOT same as Tommur CAP SOC — plug vs cap"),
    ("DWV plug (socket)", '3"', 0.589, 92, 0.4057, 0.43, 5.42, 0.0304, "NOT same as Tommur CAP SOC — plug vs cap"),
    ("DWV plug (socket)", '4"', 0.796, 56, 0.7133, 0.76, 10.50, 0.0912, "NOT same as Tommur CAP SOC — plug vs cap"),
]

# Pipe: Green 6m sticks. Convert to $/ft. Tommur sheet has no PVC pipe DDP.
PIPE = [
    ("SCH40 PVC pipe (6m)", '2"', 9.551, 3000, 19.685, "60.32mm OD × 3.91mm wall × 6m. No Tommur pipe DDP on file."),
    ("SCH40 PVC pipe (6m)", '3"', 19.836, 3000, 19.685, "88.9mm OD × 5.49mm wall × 6m."),
    ("SCH40 PVC pipe (6m)", '4"', 28.237, 3000, 19.685, "114.3mm OD × 6.02mm wall × 6m."),
    ("DWV pipe (6m)", '2"', 4.068, 3000, 19.685, "60.32mm × 1.6mm wall — foam-core/DWV thin wall."),
    ("DWV pipe (6m)", '3"', 7.527, 3000, 19.685, "88.9mm × 2.0mm wall."),
    ("DWV pipe (6m)", '4"', 10.063, 3000, 19.685, "114.3mm × 2.0mm wall."),
]


def fob_from_exw(exw):
    return round(exw * (1 + INLAND), 4)


def landed(exw, freight):
    return round(fob_from_exw(exw) * (1 + DUTY) + freight, 4)


def pct(a, b):
    if a is None or b is None or b == 0:
        return None
    return round((a - b) / b * 100, 1)


def margin(sell, cost):
    if sell is None or cost is None or sell <= 0:
        return None
    return round((sell - cost) / sell * 100, 1)


def cheaper(gv_landed, tommur_ddp):
    if gv_landed is None or tommur_ddp is None:
        return "n/a"
    if gv_landed < tommur_ddp:
        return "Green Valve landed cheaper"
    return "Tommur DDP cheaper"


def csv_escape(v):
    s = "" if v is None else str(v)
    if any(c in s for c in ',\"\n'):
        return '"' + s.replace('"', '""') + '"'
    return s


def fitting_rows():
    out = []
    for item, size, exw, qty, tfob, tddp, sell, fr, note in ROWS:
        gv_fob = fob_from_exw(exw)
        gv_l = landed(exw, fr)
        out.append(
            {
                "Item": item,
                "Size": size,
                "Green_EXW_USD": exw,
                "Green_FOB_est_USD": gv_fob,
                "Green_landed_USD": gv_l,
                "PI_qty": qty,
                "Tommur_FOB_USD": tfob,
                "Tommur_DDP_USD": tddp,
                "EXW_vs_Tommur_FOB_pct": pct(exw, tfob),
                "Landed_vs_Tommur_DDP_pct": pct(gv_l, tddp),
                "Better_buy": cheaper(gv_l, tddp),
                "Sell_USD": sell,
                "Margin_on_Tommur_DDP_pct": margin(sell, tddp),
                "Margin_on_Green_landed_pct": margin(sell, gv_l),
                "Notes": note,
            }
        )
    return out


def pipe_rows():
    out = []
    for item, size, exw_stick, qty, ft, note in PIPE:
        per_ft = round(exw_stick / ft, 4)
        landed_ft = round(fob_from_exw(per_ft) * (1 + DUTY) + 0.06, 4)
        out.append(
            {
                "Item": item,
                "Size": size,
                "Green_EXW_per_6m": exw_stick,
                "Green_EXW_per_ft": per_ft,
                "Green_landed_per_ft_est": landed_ft,
                "PI_qty_sticks": qty,
                "Line_EXW_USD": round(exw_stick * qty, 2),
                "Notes": note,
            }
        )
    return out


F_HEADERS = [
    "Item",
    "Size",
    "Green_EXW_USD",
    "Green_FOB_est_USD",
    "Green_landed_USD",
    "PI_qty",
    "Tommur_FOB_USD",
    "Tommur_DDP_USD",
    "EXW_vs_Tommur_FOB_pct",
    "Landed_vs_Tommur_DDP_pct",
    "Better_buy",
    "Sell_USD",
    "Margin_on_Tommur_DDP_pct",
    "Margin_on_Green_landed_pct",
    "Notes",
]
P_HEADERS = [
    "Item",
    "Size",
    "Green_EXW_per_6m",
    "Green_EXW_per_ft",
    "Green_landed_per_ft_est",
    "PI_qty_sticks",
    "Line_EXW_USD",
    "Notes",
]


def write_csv(rows, headers, path):
    lines = [",".join(headers)]
    for r in rows:
        lines.append(",".join(csv_escape(r[h]) for h in headers))
    path.write_text("\n".join(lines) + "\n")


def write_xlsx(frows, prows, path):
    if Workbook is None:
        return
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    green = PatternFill("solid", fgColor="D9EAD3")
    red = PatternFill("solid", fgColor="F4CCCC")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    money = {
        "Green_EXW_USD",
        "Green_FOB_est_USD",
        "Green_landed_USD",
        "Tommur_FOB_USD",
        "Tommur_DDP_USD",
        "Sell_USD",
        "Green_EXW_per_6m",
        "Green_EXW_per_ft",
        "Green_landed_per_ft_est",
        "Line_EXW_USD",
    }
    pcts = {
        "EXW_vs_Tommur_FOB_pct",
        "Landed_vs_Tommur_DDP_pct",
        "Margin_on_Tommur_DDP_pct",
        "Margin_on_Green_landed_pct",
    }

    def fill_sheet(ws, rows, headers, widths):
        for col, h in enumerate(headers, 1):
            c = ws.cell(1, col, h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="center")
        for i, r in enumerate(rows, 2):
            for col, h in enumerate(headers, 1):
                v = r[h]
                cell = ws.cell(i, col, v)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if h in money and isinstance(v, (int, float)):
                    cell.number_format = '"$"#,##0.0000'
                if h in pcts and isinstance(v, (int, float)):
                    cell.number_format = '0.0"%"'
                    if h.startswith("Margin") and v < 50:
                        cell.fill = yellow
                    if h.startswith("Margin") and v >= 70:
                        cell.fill = green
                    if h in ("EXW_vs_Tommur_FOB_pct", "Landed_vs_Tommur_DDP_pct") and v < 0:
                        cell.fill = green
                    if h in ("EXW_vs_Tommur_FOB_pct", "Landed_vs_Tommur_DDP_pct") and v > 0:
                        cell.fill = red
                if h == "Better_buy":
                    if v == "Green Valve landed cheaper":
                        cell.fill = green
                    elif v == "Tommur DDP cheaper":
                        cell.fill = red
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        ws.row_dimensions[1].height = 30
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    ws = wb.active
    ws.title = "Fittings"
    fill_sheet(
        ws,
        frows,
        F_HEADERS,
        [22, 12, 14, 14, 14, 10, 14, 14, 16, 18, 26, 12, 16, 18, 40],
    )
    wp = wb.create_sheet("Pipe")
    fill_sheet(wp, prows, P_HEADERS, [24, 10, 16, 16, 18, 14, 14, 70])

    info = wb.create_sheet("Source")
    facts = [
        ("PDF filename", "All products - Ningbo Zhenpeng Plumbing Fittings Co., Ltd_.pdf"),
        ("Actual seller", "ZHEJIANG GREEN VALVE AND FITTING CO., LTD. — NOT Zhenpeng"),
        ("PI", "YMF26060303X dated 2026/6/3, buyer Baruch Grossman"),
        ("Incoterm", "EXW our factory (Taizhou). Not FOB. Not DDP."),
        ("Addr", "NO.525 Tengyun Road, Jiaojiang, Taizhou, Zhejiang 318000"),
        ("Tel", "0086-576-88123188 / 15867653812 / 15867628609"),
        ("Bank", "Agricultural Bank of China Zhejiang Branch 19911014040002072 SWIFT ABOCCNBJ110"),
        ("PI total", "EXW $240,505.87 — almost all pipe ($237.8k). Fittings ~$2.7k."),
        ("Payment", "30% deposit, 70% before ship, 45 days after deposit"),
        ("Quality", "Subject to seller samples"),
        ("Landed method", "EXW × 1.03 inland ≈ FOB, then × 1.428 duty (5.3+25+12.5) + Tommur-sheet freight/pc"),
        ("Zhenpeng", "This PI has zero PEX / F2159. Still email Zhenpeng separately for poly-alloy."),
    ]
    info["A1"] = "Field"
    info["B1"] = "Value"
    info["A1"].fill = header_fill
    info["B1"].fill = header_fill
    info["A1"].font = header_font
    info["B1"].font = header_font
    for i, (k, v) in enumerate(facts, 2):
        info.cell(i, 1, k)
        info.cell(i, 2, v)
        info.cell(i, 2).alignment = Alignment(wrap_text=True)
    info.column_dimensions["A"].width = 16
    info.column_dimensions["B"].width = 110
    wb.save(path)


def main():
    frows = fitting_rows()
    prows = pipe_rows()
    write_csv(frows, F_HEADERS, OUT / "GreenValve_vs_Tommur.csv")
    write_xlsx(frows, prows, OUT / "GreenValve_vs_Tommur.xlsx")
    print("FITTINGS  EXW vs Tommur FOB | landed vs Tommur DDP")
    for r in frows:
        if "NOT same" in r["Notes"]:
            continue
        print(
            f"{r['Item'][:18]:18} {r['Size']:8}  "
            f"GV EXW {r['Green_EXW_USD']:.3f}  T FOB {r['Tommur_FOB_USD']:.3f} ({r['EXW_vs_Tommur_FOB_pct']:+.0f}%)  "
            f"GV land {r['Green_landed_USD']:.3f}  T DDP {r['Tommur_DDP_USD']:.3f} ({r['Landed_vs_Tommur_DDP_pct']:+.0f}%)  "
            f"{r['Better_buy']}"
        )
    print("\nPIPE")
    for r in prows:
        print(
            f"{r['Item']:22} {r['Size']:4}  EXW ${r['Green_EXW_per_ft']:.4f}/ft  "
            f"landed ~${r['Green_landed_per_ft_est']:.4f}/ft  line ${r['Line_EXW_USD']:,.0f}"
        )


if __name__ == "__main__":
    main()
