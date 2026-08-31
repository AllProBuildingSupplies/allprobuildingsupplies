#!/usr/bin/env python3
"""Build a simple competitive margin sheet: FOB, Landed, Online, Margin$, Margin%."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
COST_XLSX = HERE / "Tommur_Cost_Margin_Tracker.xlsx"
OUT_XLSX = HERE / "Competitive_Margin_Sheet.xlsx"
OUT_CSV = HERE / "Competitive_Margin_Sheet.csv"
ART_XLSX = Path("/opt/cursor/artifacts/Competitive_Margin_Sheet.xlsx")
ART_CSV = Path("/opt/cursor/artifacts/Competitive_Margin_Sheet.csv")


def norm_size(s: str | None) -> str:
    if s is None:
        return ""
    t = str(s).strip().upper()
    t = t.replace("″", '"').replace("”", '"').replace("“", '"')
    t = t.replace("×", "X").replace("x", "X")
    t = re.sub(r"\s+", "", t)
    # strip odd width annotations e.g. 1/2(21.34x2.77mm）
    t = re.sub(r"\(.*?\)", "", t)
    # also strip fullwidth / leftover paren garbage
    t = re.sub(r"[（(].*$", "", t)
    t = t.replace("\"", "")
    # unicode fractions — compound forms first
    t = (
        t.replace("1½", "1-1/2")
        .replace("1¼", "1-1/4")
        .replace("2½", "2-1/2")
        .replace("¾", "3/4")
        .replace("½", "1/2")
        .replace("¼", "1/4")
    )
    # Pipe labels like 1X2.6 or 1/2X1.73 (nominal x wall mm/in) → nominal only.
    # Do NOT strip reducing fittings (6X3, 2X1-1/2, 10X10X4, etc.).
    if "X" in t:
        parts = t.split("X")
        if len(parts) == 2 and re.match(r"^[\d\-/]+$", parts[0]) and re.match(r"^\d+\.\d+$", parts[1]):
            t = parts[0]
    return t


def money(x):
    if x is None or x == "":
        return None
    return round(float(x), 2)


def pct(margin, online):
    if online is None or online == 0 or margin is None:
        return None
    return round(margin / online * 100, 1)


def load_master():
    wb = openpyxl.load_workbook(COST_XLSX, data_only=True)
    ws = wb["Cost_Margin_Master"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c]: ws.cell(r, c + 1).value for c in range(len(headers))}
        if row.get("FOB_USD") in (None, ""):
            continue
        landed = row.get("Est_Landed_per_Pc_45HQ") or row.get("Est_Landed_per_Pc_40ft")
        landed_note = "45HQ" if row.get("Est_Landed_per_Pc_45HQ") else (
            "40ft" if row.get("Est_Landed_per_Pc_40ft") else "ex-freight"
        )
        if landed in (None, ""):
            duty = row.get("Duty_Tariff_Only_Pct") or 0.428
            landed = float(row["FOB_USD"]) * (1 + float(duty))
            landed_note = "ex-freight"
        unit = "ft" if (row.get("Sell_Unit") or "").lower().find("ft") >= 0 else "pc"
        rows.append(
            {
                "code": row.get("APBS_Item_Code") or "",
                "tommur": row.get("Tommur_Code") or "",
                "material": row.get("Material") or "",
                "desc": (row.get("Description") or "").strip(),
                "size": row.get("Size") or "",
                "size_n": norm_size(row.get("Size")),
                "unit": unit,
                "fob": float(row["FOB_USD"]),
                "landed": float(landed),
                "landed_note": landed_note,
                "on_web": row.get("On_Website") or "",
                "sell": row.get("Selling_Price"),
            }
        )
    return rows


def load_price_maps():
    copper = json.loads(Path("/opt/cursor/artifacts/copper_price_survey.json").read_text())
    pvc = json.loads(Path("/tmp/prices_pvc.json").read_text())
    other = json.loads(Path("/tmp/prices_cpvc_pex.json").read_text())

    def put(m, key, price, source, notes=""):
        if price is None:
            return
        k = (key[0], norm_size(key[1]))
        prev = m.get(k)
        if prev is None or price < prev["price"]:
            m[k] = {"price": float(price), "source": source or "", "notes": notes or ""}

    cmap = {}
    for r in copper:
        fam = r["family"]
        put(cmap, (fam, r["size"]), r.get("cheapest_price_usd"), r.get("source"), r.get("notes") or "")

    pmap = {}
    for r in pvc:
        put(pmap, (r["item"], r["size"]), r.get("cheapest_price_usd"), r.get("source_site"), r.get("notes") or "")

    omap = {}
    for r in other:
        put(omap, (r["family"], r["size"]), r.get("cheapest_price_usd"), r.get("source"), "")

    return cmap, pmap, omap


def match_online(row, cmap, pmap, omap):
    mat = row["material"]
    desc = row["desc"].upper()
    sz = row["size_n"]

    def from_map(m, family):
        hit = m.get((family, sz))
        if hit:
            return hit["price"], hit["source"], hit["notes"]
        # try with quotes stripped variants already in norm
        return None, None, None

    # --- Copper ---
    if mat == "Copper":
        if "PIPE L" in desc or desc == "COPPER PIPE L":
            return from_map(cmap, "COPPER PIPE L")
        if "PIPE K" in desc or desc == "COPPER PIPE K":
            return from_map(cmap, "COPPER PIPE K")
        if "ELBOW 90" in desc:
            return from_map(cmap, "ELBOW 90")
        if desc.startswith("TEE") and "RED" not in desc:
            return from_map(cmap, "TEE")
        if "RED TEE" in desc:
            return from_map(cmap, "RED TEE")
        if "COUPLING" in desc:
            return from_map(cmap, "COUPLING")
        if "REDUCER" in desc:
            return from_map(cmap, "REDUCER")
        if "ADAPTER" in desc and "PEX" in desc:
            # Cheap brass PEX×sweat adapters are not comparable to copper-to-PEX adapters
            p, s, n = from_map(cmap, "ADAPTER copper to PEX")
            if p is not None and p < 5 and row["fob"] > 2:
                return None, None, "online brass PEX adapter not comparable to copper adapter"
            return p, s, n
        if "STUB" in desc:
            return from_map(cmap, "COPPER ELBOW STUB OUT")
        if "DIELEC" in desc or "UNION" in desc:
            return from_map(cmap, "DIELECTRIC UNION")

    # --- PVC ---
    if mat == "PVC":
        if "D1785" in desc or "SCH40 PVC" in desc or "SCH 40" in desc:
            return from_map(pmap, "PVC SCH40 PIPE")
        if "1/4 BEND" in desc:
            return from_map(pmap, "1/4 BEND 90 HxH")
        if "1/8 BEND" in desc:
            return from_map(pmap, "1/8 BEND 45 HxH")
        if "45" in desc and "ELL" in desc:
            return from_map(pmap, "45 ELL SOC x SOC")
        if "REDUCING SANITARY" in desc:
            return from_map(pmap, "REDUCING SANITARY TEE")
        if "SANITARY TEE" in desc:
            return from_map(pmap, "SANITARY TEE HxHxH")
        if "REDUCING WYE" in desc:
            return from_map(pmap, "REDUCING WYE")
        if desc.startswith("WYE"):
            return from_map(pmap, "WYE HxHxH")
        if "INCREASER" in desc or "REDUCER" in desc:
            return from_map(pmap, "PIPE INCREASER-REDUCER HxH")
        if "P-TRAP" in desc or "P TRAP" in desc:
            return from_map(pmap, "P-TRAP")

    # --- CPVC ---
    if mat == "CPVC":
        tom = (row["tommur"] or "").upper()
        if "SCH80" in desc or "SCH 80" in desc:
            return from_map(omap, "CPVC_SCH80_pipe")
        if "SDR-11" in desc or "SDR11" in desc:
            return from_map(omap, "CPVC_SDR11_pipe")
        if "SDR-13.5" in desc or "SDR13.5" in desc:
            # rarely stocked; leave blank
            return None, None, "SDR-13.5 not commonly stocked online"
        if "COUPLING" in desc:
            if "SCH 80" in tom or "SCH80" in tom:
                return from_map(omap, "CPVC_SCH80_coupling")
            if "2846" in tom:
                p, s, n = from_map(omap, "CPVC_CTS_coupling")
                if p is not None:
                    return p, s, n
                return from_map(omap, "CPVC_SCH80_coupling")
            return from_map(omap, "CPVC_SCH80_coupling")
        if "TEE" in desc:
            if "SCH 80" in tom or "SCH80" in tom:
                return from_map(omap, "CPVC_SCH80_tee")
            if "2846" in tom:
                p, s, n = from_map(omap, "CPVC_CTS_tee")
                if p is not None:
                    return p, s, n
                return from_map(omap, "CPVC_SCH80_tee")
            return from_map(omap, "CPVC_SCH80_tee")

    # --- PEX ---
    if mat == "PEX":
        if "PIPE" in desc:
            return from_map(omap, "PEX-B_pipe")
        if "ELBOW" in desc:
            return from_map(omap, "PEX_elbow")
        if "REDUCER" in desc:
            return from_map(omap, "PEX_reducer")

    # --- Insulation ---
    if mat == "INSULATION" or "INSULATION" in desc:
        return from_map(omap, "insulation")

    return None, None, None


def build():
    rows = load_master()
    cmap, pmap, omap = load_price_maps()

    out = []
    for r in rows:
        online, source, notes = match_online(r, cmap, pmap, omap)
        # Zero FOB rows are placeholders — do not treat as real cost
        if r["fob"] == 0:
            online, source = None, None
            notes = ((notes + "; ") if notes else "") + "FOB is $0 in tracker — excluded until real FOB entered"
        margin_usd = None if online is None else round(online - r["landed"], 2)
        margin_pct = pct(margin_usd, online)
        out.append(
            {
                "Code": r["code"],
                "Material": r["material"],
                "Description": r["desc"],
                "Size": r["size"],
                "Unit": r["unit"],
                "On_Website": r["on_web"],
                "FOB": money(r["fob"]),
                "Landed": money(r["landed"]),
                "Landed_Basis": r["landed_note"],
                "Cheapest_Online": money(online) if online is not None else None,
                "Online_Source": source or "",
                "Margin_USD": margin_usd,
                "Margin_Pct": margin_pct,
                "Notes": notes or "",
            }
        )

    # Deduplicate identical CPVC SCH80 vs 2846 rows that share same FOB/size/desc when both match same online
    # Keep both but sort sensibly
    out.sort(key=lambda x: (x["Material"], x["Description"], x["Size"], x["Code"] or "", x.get("Notes") or ""))

    # Write CSV
    import csv

    fields = [
        "Code",
        "Material",
        "Description",
        "Size",
        "Unit",
        "On_Website",
        "FOB",
        "Landed",
        "Cheapest_Online",
        "Online_Source",
        "Margin_USD",
        "Margin_Pct",
        "Landed_Basis",
        "Notes",
    ]
    for path in (OUT_CSV, ART_CSV):
        try:
            with open(path, "w", newline="") as f:
                w = csv.DictWriter(f, fieldnames=fields)
                w.writeheader()
                for row in out:
                    w.writerow({k: row.get(k) for k in fields})
        except OSError as e:
            print(f"WARN write {path}: {e}")

    # Simple Excel: one sheet, clean columns user asked for + identity
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Margins"

    # User-requested core + minimal identity
    headers = [
        "Code",
        "Material",
        "Description",
        "Size",
        "Unit",
        "FOB",
        "Landed",
        "Cheapest Online",
        "Online Source",
        "Margin $",
        "Margin %",
        "On Website",
        "Notes",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="F2F2F2")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(out, 2):
        vals = [
            row["Code"],
            row["Material"],
            row["Description"],
            row["Size"],
            row["Unit"],
            row["FOB"],
            row["Landed"],
            row["Cheapest_Online"],
            row["Online_Source"],
            row["Margin_USD"],
            row["Margin_Pct"],
            row["On_Website"],
            row["Notes"] if row["Notes"] else (f"landed={row['Landed_Basis']}" if row["Landed_Basis"] != "45HQ" else ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v)
            cell.border = thin
            if c in (6, 7, 8, 10):
                cell.number_format = '$#,##0.00'
            if c == 11 and v is not None:
                cell.number_format = '0.0"%"'
                # store as percent number? User wants e.g. 45.2 meaning 45.2% — keep as number with % suffix format
                # openpyxl percent format expects fraction; keep raw 45.2 with custom format
                cell.number_format = '0.0'

        # color margin row
        m = row["Margin_Pct"]
        online = row["Cheapest_Online"]
        fill = None
        if online is None:
            fill = gray
        elif m is not None and m >= 30:
            fill = green
        elif m is not None and m >= 15:
            fill = yellow
        elif m is not None and m < 0:
            fill = red
        if fill:
            for c in (10, 11):
                ws.cell(i, c).fill = fill

    # widths
    widths = [16, 10, 36, 14, 6, 10, 10, 14, 16, 10, 10, 10, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:M{len(out)+1}"
    ws.freeze_panes = "A2"

    # Summary sheet
    sm = wb.create_sheet("Summary", 0)
    priced = [r for r in out if r["Cheapest_Online"] is not None]
    neg = [r for r in priced if (r["Margin_Pct"] or 0) < 0]
    low = [r for r in priced if 0 <= (r["Margin_Pct"] or 0) < 15]
    mid = [r for r in priced if 15 <= (r["Margin_Pct"] or 0) < 30]
    high = [r for r in priced if (r["Margin_Pct"] or 0) >= 30]
    missing = [r for r in out if r["Cheapest_Online"] is None]

    summary_lines = [
        ("Competitive Margin Sheet", ""),
        ("Generated from Tommur FOB/landed + online price survey", ""),
        ("", ""),
        ("SKUs with FOB", len(out)),
        ("With online price found", len(priced)),
        ("Online price not found", len(missing)),
        ("", ""),
        ("Margin vs cheapest online (at landed cost)", ""),
        ("Negative margin (underwater)", len(neg)),
        ("0–15% margin", len(low)),
        ("15–30% margin", len(mid)),
        ("30%+ margin", len(high)),
        ("", ""),
        ("Rules", ""),
        ("Pipe FOB/Landed/Online", "per foot"),
        ("Everything else", "per piece"),
        ("Landed", "45'HQ freight when dims exist; else FOB×(1+duty) ex-freight"),
        ("Online", "Cheapest bulk/each unit price found (SupplyHouse/HD/PFO/FlexPVC/PexUniverse/etc.)"),
        ("Margin $", "Cheapest Online − Landed"),
        ("Margin %", "(Online − Landed) / Online"),
        ("", ""),
        ("IMPORTANT CAVEATS", ""),
        ("Copper pipe", "Yellow FOB labeled per-ft but sits ABOVE US retail $/ft (e.g. 1/2\" L FOB $6.93 vs HD ~$2.66/ft). Verify with factory whether yellow FOB is $/ft, $/stick, or other unit before pricing copper pipe."),
        ("Insulation", "Compared to commodity PE foam online. Your FOB is much higher — may be thicker/rubber product or different unit; treat negative margins as a product-match warning."),
        ("CPVC 2846 CTS fittings ≤1\"", "US big-box CTS pro-packs are extremely cheap; import landed often loses to HD CTS. SCH80 line has healthy margins vs SCH80 online."),
        ("Zero FOB rows", "Excluded from online margin (placeholder $0 in tracker)."),
        ("", ""),
        ("Color key on Margins sheet", ""),
        ("Green", "≥30% margin"),
        ("Yellow", "15–30%"),
        ("Red", "negative"),
        ("Gray", "no online price found"),
    ]
    sm["A1"].font = Font(bold=True, size=14)
    for i, (a, b) in enumerate(summary_lines, 1):
        sm.cell(i, 1, a)
        sm.cell(i, 2, b)
        if i == 1:
            sm.cell(i, 1).font = Font(bold=True, size=14)
    sm.column_dimensions["A"].width = 48
    sm.column_dimensions["B"].width = 80

    # Notes sheet
    notes = wb.create_sheet("Research_Notes")
    notes_text = [
        "Sources checked: supplyhouse.com (often cheapest on box/bulk; Cloudflare blocked automated fetch — used indexed/cached prices where available), homedepot.com, pvcfittingsonline.com, flexpvc.com, pexuniverse.com, plumbingsell.com, procuru.com, acandb.com, zoro.com, fwwebb.com.",
        "Bulk/box unit price used when lower than single-piece list.",
        "CPVC SCH80 fittings compared to SCH80 online; CPVC 2846 SCH40/CTS compared to CTS FlowGuard-style where sized ≤1\".",
        "CPVC SDR-13.5 pipe is rarely stocked at US online retailers — left blank.",
        "Insulation online comps are commodity polyethylene foam; if your Tommur insulation is a thicker/rubber product, comps may be too low.",
        "Copper pipe HD prices can be ZIP-local; used advertised bulk (buy 10+) when available.",
        "This is a pricing decision aid, not a customs/broker quote. Landed costs come from the prior Tommur_Cost_Margin_Tracker assumptions.",
        "Only SKUs with yellow-column FOB in the Tommur tracker are included (205). Website SKUs still missing FOB are listed in that tracker Data_Gaps sheet.",
    ]
    for i, t in enumerate(notes_text, 1):
        notes.cell(i, 1, t)
        notes.row_dimensions[i].height = 45
    notes.column_dimensions["A"].width = 120
    notes["A1"].alignment = Alignment(wrap_text=True)

    for path in (OUT_XLSX, ART_XLSX):
        try:
            wb.save(path)
        except OSError as e:
            print(f"WARN save {path}: {e}")

    print(f"Wrote {OUT_XLSX} and {OUT_CSV}")
    print(f"rows={len(out)} priced={len(priced)} neg={len(neg)} low={len(low)} mid={len(mid)} high={len(high)} missing={len(missing)}")
    # show a few underwater and top margin
    priced_sorted = sorted(priced, key=lambda r: r["Margin_Pct"] if r["Margin_Pct"] is not None else -999)
    print("\nWorst margins:")
    for r in priced_sorted[:8]:
        print(f"  {r['Material']} {r['Description']} {r['Size']}: online={r['Cheapest_Online']} landed={r['Landed']} m%={r['Margin_Pct']} src={r['Online_Source']}")
    print("\nBest margins:")
    for r in priced_sorted[-8:]:
        print(f"  {r['Material']} {r['Description']} {r['Size']}: online={r['Cheapest_Online']} landed={r['Landed']} m%={r['Margin_Pct']} src={r['Online_Source']}")


if __name__ == "__main__":
    build()
