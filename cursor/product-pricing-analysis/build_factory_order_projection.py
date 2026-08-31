#!/usr/bin/env python3
"""
Factory order projection: 3-month supply of high-margin sellers,
net of on-hand + inbound containers 3/4 − open backorders, with container CBM fill.
"""

from __future__ import annotations

import csv
import json
import math
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
ORDERS_JSON = Path("/tmp/plan_orders.json")
PRODUCTS_JSON = Path("/tmp/plan_products.json")
BACKORDERS_JSON = Path("/tmp/plan_backorders.json")
INBOUND_JSON = REPO / "data" / "inbound-containers.json"
MARGINS_CSV = HERE / "Competitive_Margin_Sheet.csv"
COST_XLSX = HERE / "Tommur_Cost_Margin_Tracker.xlsx"
CATALOG_CSV = REPO / "assets" / "products.csv"

OUT_XLSX = HERE / "Factory_Order_3Mo_Container.xlsx"
OUT_CSV = HERE / "Factory_Order_For_Tommur.csv"
ART_XLSX = Path("/opt/cursor/artifacts/Factory_Order_3Mo_Container.xlsx")
ART_CSV = Path("/opt/cursor/artifacts/Factory_Order_For_Tommur.csv")

CBM_40 = 67.7
CBM_45HQ = 86.0
HIGH_MARGIN_PCT = 30.0  # "make a lot of money"
MONTHS_SUPPLY = 3.0


def norm_size(s) -> str:
    t = str(s or "").strip().replace('"', "").replace("″", "").replace("”", "")
    t = t.replace("×", "x").replace("X", "x")
    t = t.replace(" ", "")
    return t


def ceil_pack(qty: float, pack: int) -> int:
    if qty <= 0:
        return 0
    pack = max(int(pack or 1), 1)
    return int(math.ceil(qty / pack) * pack)


def load_all():
    orders = json.loads(ORDERS_JSON.read_text())["orders"]
    items = json.loads(ORDERS_JSON.read_text())["items"]
    products = json.loads(PRODUCTS_JSON.read_text())["products"]
    backorders = json.loads(BACKORDERS_JSON.read_text())["backorders"]
    inbound_doc = json.loads(INBOUND_JSON.read_text())

    order_ids = {o["id"] for o in orders}
    sales = defaultdict(lambda: {"qty": 0, "rev": 0.0, "orders": set()})
    for it in items:
        if it["order_id"] not in order_ids:
            continue
        key = (it["product_sku"], norm_size(it["size"]))
        sales[key]["qty"] += it["quantity"] or 0
        sales[key]["rev"] += (it["quantity"] or 0) * (it["price_at_purchase"] or 0)
        sales[key]["orders"].add(it["order_id"])

    on_hand = {}
    meta = {}
    for p in products:
        key = (p["code"], norm_size(p["size"]))
        on_hand[key] = float(p.get("qty") or 0)
        meta[key] = {
            "description": p.get("description") or "",
            "material": p.get("material") or "",
            "pack": int(p.get("pack") or 1) or 1,
            "price": float(p.get("price") or 0),
            "tommur": p.get("tommur_code") or "",
            "lesso": p.get("lesso_code") or "",
        }

    # catalog packs / tommur fallback
    for r in csv.DictReader(open(CATALOG_CSV)):
        key = (r["Code"], norm_size(r["Size"]))
        if key not in meta:
            meta[key] = {
                "description": r.get("Description") or "",
                "material": r.get("Material") or "",
                "pack": int(float(r.get("Pack") or 1)) or 1,
                "price": float(r.get("Price") or 0),
                "tommur": r.get("Tommur-Code") or "",
                "lesso": r.get("Lesso-Code") or "",
            }
        else:
            if not meta[key]["tommur"]:
                meta[key]["tommur"] = r.get("Tommur-Code") or ""
            if not meta[key]["pack"] or meta[key]["pack"] == 1:
                pk = int(float(r.get("Pack") or 1)) or 1
                if pk > 1:
                    meta[key]["pack"] = pk

    inbound = defaultdict(float)
    inbound_by_c = {"3": defaultdict(float), "4": defaultdict(float)}
    for c in inbound_doc["containers"]:
        which = "3" if "3" in (c.get("label") or "") else "4"
        for it in c["items"]:
            key = (it["code"], norm_size(it["size"]))
            inbound[key] += it["qty"] or 0
            inbound_by_c[which][key] += it["qty"] or 0

    bo = defaultdict(float)
    for b in backorders:
        key = (b["code"], norm_size(b["size"]))
        bo[key] += b["qty_backordered"] or 0

    # margins
    margins = {}
    for r in csv.DictReader(open(MARGINS_CSV)):
        code = (r.get("Code") or "").strip()
        size = norm_size(r.get("Size"))
        if not code:
            continue
        try:
            mp = float(r["Margin_Pct"]) if r.get("Margin_Pct") not in (None, "") else None
        except ValueError:
            mp = None
        try:
            landed = float(r["Landed"]) if r.get("Landed") not in (None, "") else None
        except ValueError:
            landed = None
        try:
            online = float(r["Cheapest_Online"]) if r.get("Cheapest_Online") not in (None, "") else None
        except ValueError:
            online = None
        margins[(code, size)] = {
            "margin_pct": mp,
            "landed": landed,
            "online": online,
            "fob": float(r["FOB"]) if r.get("FOB") not in (None, "") else None,
            "source": r.get("Online_Source") or "",
        }

    # cost / CBM
    wb = openpyxl.load_workbook(COST_XLSX, data_only=True)
    ws = wb["Cost_Margin_Master"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    cost = {}
    for r in range(2, ws.max_row + 1):
        row = {headers[i]: ws.cell(r, i + 1).value for i in range(len(headers))}
        code = (row.get("APBS_Item_Code") or "").strip()
        size = norm_size(row.get("Size"))
        if not code:
            continue
        cost[(code, size)] = row
        if code in meta or True:
            # enrich tommur from cost sheet
            t = row.get("Tommur_Code") or ""
            if (code, size) in meta and not meta[(code, size)]["tommur"] and t:
                meta[(code, size)]["tommur"] = t
            pcs = row.get("Pcs_per_Carton")
            if pcs and (code, size) in meta:
                if meta[(code, size)]["pack"] <= 1:
                    meta[(code, size)]["pack"] = int(pcs)

    dates = [
        datetime.fromisoformat(o["created_at"].replace("Z", "+00:00"))
        for o in orders
        if o.get("created_at")
    ]
    span_days = max((max(dates) - min(dates)).days, 1) if dates else 1
    span_months = span_days / 30.44

    return {
        "orders": orders,
        "sales": sales,
        "on_hand": on_hand,
        "meta": meta,
        "inbound": inbound,
        "inbound_by_c": inbound_by_c,
        "bo": bo,
        "margins": margins,
        "cost": cost,
        "span_months": span_months,
        "span_days": span_days,
        "dates": dates,
    }


def is_money_maker(code: str, size: str, material: str, margins: dict) -> tuple[bool, str]:
    m = margins.get((code, size))
    if m and m.get("margin_pct") is not None and m["margin_pct"] >= HIGH_MARGIN_PCT:
        return True, f"margin {m['margin_pct']:.0f}%"
    # PVC DWV fittings (proven family) — siblings with FOB are 70–90%
    if (material or "").upper() == "PVC" and not any(
        x in code.upper() for x in ("PIPE-FOAM", "PIPE-SOLID", "PIPE-INSLTN", "INSLTN")
    ):
        return True, "PVC DWV fitting family (peer margins 70–90%)"
    return False, ""


def build_rows(data):
    sales = data["sales"]
    on_hand = data["on_hand"]
    meta = data["meta"]
    inbound = data["inbound"]
    bo = data["bo"]
    margins = data["margins"]
    cost = data["cost"]
    span_months = data["span_months"]

    keys = set(sales) | set(bo) | set(k for k, v in inbound.items() if v > 0)
    # also high-margin catalog items with FOB even if no sales? keep focused on demand first
    rows = []
    for key in sorted(keys):
        code, size = key
        m = meta.get(key) or {
            "description": "",
            "material": "",
            "pack": 1,
            "price": 0,
            "tommur": "",
            "lesso": "",
        }
        money, money_why = is_money_maker(code, size, m["material"], margins)
        sold = sales.get(key, {}).get("qty", 0)
        monthly = sold / span_months if span_months else 0
        need_3mo = monthly * MONTHS_SUPPLY
        oh = on_hand.get(key, 0)
        inn = inbound.get(key, 0)
        back = bo.get(key, 0)
        # After C3/C4 arrive and open BOs ship:
        available = oh + inn - back
        # Remaining BO not covered by inbound+onhand (must still order)
        bo_gap = max(0.0, back - oh - inn)
        # Forward 3-mo stock need beyond post-arrival available
        forward_gap = max(0.0, need_3mo - max(0.0, available))
        raw_order = bo_gap + forward_gap
        pack = m["pack"] or 1
        # Prefer cost-sheet pcs/ctn when available
        c = cost.get(key) or {}
        if c.get("Pcs_per_Carton"):
            pack = int(c["Pcs_per_Carton"]) or pack
        order_pcs = ceil_pack(raw_order, pack)
        cartons = int(order_pcs / pack) if pack else 0
        cbm_pc = c.get("CBM_per_Pc")
        cbm = (cbm_pc or 0) * order_pcs if order_pcs else 0
        mg = margins.get(key) or {}
        rows.append(
            {
                "code": code,
                "size": size,
                "description": m["description"],
                "material": m["material"],
                "tommur": m["tommur"] or (c.get("Tommur_Code") or ""),
                "lesso": m["lesso"] or (c.get("Lesso_Code") or ""),
                "unit": "ft" if "PIPE" in code.upper() and "INSLTN" not in code.upper() else "pc",
                "pack": pack,
                "money_maker": money,
                "money_why": money_why,
                "sold_hist": sold,
                "monthly": round(monthly, 1),
                "need_3mo": round(need_3mo, 1),
                "on_hand": oh,
                "inbound": inn,
                "backorder": back,
                "available_after": round(available, 1),
                "bo_gap": round(bo_gap, 1),
                "order_pcs": order_pcs,
                "cartons": cartons,
                "cbm": round(cbm, 4) if cbm else (None if order_pcs and not cbm_pc else round(cbm, 4)),
                "cbm_known": bool(cbm_pc),
                "margin_pct": mg.get("margin_pct"),
                "landed": mg.get("landed"),
                "online": mg.get("online"),
                "fob": mg.get("fob") or c.get("FOB_USD"),
                "sell_price": m["price"],
            }
        )
    return rows


def write_outputs(data, rows):
    money_rows = [r for r in rows if r["money_maker"] and r["order_pcs"] > 0]
    money_rows.sort(key=lambda r: (-(r["margin_pct"] or 0), -r["order_pcs"]))

    all_order = [r for r in rows if r["order_pcs"] > 0]
    cbm_known = sum(r["cbm"] or 0 for r in money_rows if r["cbm_known"])
    cbm_unknown_pcs = sum(r["order_pcs"] for r in money_rows if not r["cbm_known"])
    cartons = sum(r["cartons"] for r in money_rows)
    pcs = sum(r["order_pcs"] for r in money_rows)

    fill40 = cbm_known / CBM_40 * 100
    fill45 = cbm_known / CBM_45HQ * 100

    # CSV for factory (Tommur)
    fields = [
        "Tommur_Code",
        "APBS_Code",
        "Description",
        "Size",
        "Unit",
        "Order_Pcs",
        "Pcs_per_Carton",
        "Cartons",
        "CBM",
        "FOB_USD",
        "Notes",
    ]
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in money_rows:
            w.writerow(
                {
                    "Tommur_Code": r["tommur"],
                    "APBS_Code": r["code"],
                    "Description": r["description"],
                    "Size": r["size"],
                    "Unit": r["unit"],
                    "Order_Pcs": r["order_pcs"],
                    "Pcs_per_Carton": r["pack"],
                    "Cartons": r["cartons"],
                    "CBM": r["cbm"] if r["cbm_known"] else "",
                    "FOB_USD": round(r["fob"], 4) if r["fob"] else "",
                    "Notes": r["money_why"],
                }
            )
    try:
        ART_CSV.write_text(OUT_CSV.read_text())
    except OSError:
        pass

    # Excel workbook
    wb = openpyxl.Workbook()
    thin = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(color="FFFFFF", bold=True)
    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")

    # --- Summary ---
    sm = wb.active
    sm.title = "Summary"
    orders = data["orders"]
    dates = data["dates"]
    lines = [
        ("Factory Order Projection — 3 months high-margin supply", ""),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("SALES BASIS", ""),
        ("Non-cancelled orders", len(orders)),
        ("Date range", f"{min(dates).date()} → {max(dates).date()}" if dates else ""),
        ("Span", f"{data['span_days']} days (~{data['span_months']:.2f} months)"),
        ("Primary customer", "NJPD Plumbing & Heating LLC (all 5 orders)"),
        ("Note", "Velocity is thin (5 orders); Aug mega-order APBS-000005 (~46k pcs) dominates. Treat as directional."),
        ("", ""),
        ("INBOUND / BACKORDERS", ""),
        ("Containers", "3 (WHSU9010053) + 4 (WHSU9004718), ETA 2026-09-15"),
        ("Inbound pcs", int(sum(data["inbound"].values()))),
        ("Open backorder pcs", int(sum(data["bo"].values()))),
        ("BO lines fully covered by C3+C4", ""),
        ("", ""),
        ("ORDER RULE", ""),
        ("Money-makers", f"Margin ≥{HIGH_MARGIN_PCT:.0f}% OR PVC DWV fittings (peer margins 70–90%)"),
        ("Target", f"{MONTHS_SUPPLY:.0f} months of historical monthly sales"),
        ("Netting", "on-hand + inbound C3/C4 − open backorders, then buy the gap; also buy any BO shortfall C3/C4 won't cover"),
        ("Rounding", "Up to factory carton (pcs/ctn)"),
        ("", ""),
        ("FACTORY ORDER (money-makers only)", ""),
        ("SKU lines", len(money_rows)),
        ("Total pcs", pcs),
        ("Total cartons", cartons),
        ("CBM (known dims)", round(cbm_known, 2)),
        ("Pcs missing CBM dims", cbm_unknown_pcs),
        ("40' fill (67.7 CBM)", f"{fill40:.1f}% of known-CBM portion"),
        ("45'HQ fill (86 CBM)", f"{fill45:.1f}% of known-CBM portion"),
        ("Fits one container?", "YES — under both 40' and 45'HQ on known CBM; room to top-up" if fill40 < 95 else "TIGHT / OVER — review"),
        ("", ""),
        ("RECOMMENDATION", ""),
        (
            "Container fill",
            f"Core 3-mo money-maker order uses ~{fill45:.0f}% of a 45'HQ (known CBM). "
            "Use remaining cube to (a) deepen top sellers another 1–2 months, or (b) add high-margin CPVC SCH80 pipe/fittings that have FOB but no sales yet.",
        ),
    ]
    # BO coverage stats
    full = sum(1 for k, v in data["bo"].items() if data["on_hand"].get(k, 0) + data["inbound"].get(k, 0) >= v)
    partial = sum(
        1
        for k, v in data["bo"].items()
        if 0 < data["on_hand"].get(k, 0) + data["inbound"].get(k, 0) < v
    )
    none = sum(1 for k, v in data["bo"].items() if data["on_hand"].get(k, 0) + data["inbound"].get(k, 0) <= 0)
    lines[14] = ("BO lines covered / partial / none", f"{full} / {partial} / {none}")

    for i, (a, b) in enumerate(lines, 1):
        sm.cell(i, 1, a)
        sm.cell(i, 2, b)
        if i == 1:
            sm.cell(i, 1).font = Font(bold=True, size=14)
    sm.column_dimensions["A"].width = 36
    sm.column_dimensions["B"].width = 100

    def sheet_from_rows(title, rowset, cols):
        ws = wb.create_sheet(title)
        for c, h in enumerate(cols, 1):
            cell = ws.cell(1, c, h)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.border = thin
        for i, r in enumerate(rowset, 2):
            vals = [
                r["code"],
                r["tommur"],
                r["description"],
                r["size"],
                r["unit"],
                r["pack"],
                r["sold_hist"],
                r["monthly"],
                r["need_3mo"],
                r["on_hand"],
                r["inbound"],
                r["backorder"],
                r["available_after"],
                r["order_pcs"],
                r["cartons"],
                r["cbm"] if r["cbm_known"] else "",
                r["margin_pct"],
                r["fob"],
                r["landed"],
                r["online"],
                r["money_why"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(i, c, v)
                cell.border = thin
            if r["order_pcs"] > 0:
                ws.cell(i, 14).fill = green
            if r["available_after"] < 0:
                ws.cell(i, 13).fill = red
            if r["margin_pct"] is not None and r["margin_pct"] >= 30:
                ws.cell(i, 17).fill = green
            elif r["margin_pct"] is not None and r["margin_pct"] < 0:
                ws.cell(i, 17).fill = red
        ws.auto_filter.ref = f"A1:U{len(rowset)+1}"
        ws.freeze_panes = "A2"
        widths = [14, 18, 28, 12, 6, 8, 10, 10, 10, 10, 10, 10, 12, 10, 8, 8, 10, 8, 8, 8, 36]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    cols = [
        "Code",
        "Tommur",
        "Description",
        "Size",
        "Unit",
        "Pack",
        "Sold_Hist",
        "Monthly",
        "Need_3Mo",
        "On_Hand",
        "Inbound_C3C4",
        "Backorder",
        "Avail_After_C3C4",
        "Order_Pcs",
        "Cartons",
        "CBM",
        "Margin_%",
        "FOB",
        "Landed",
        "Online",
        "Why_Money_Maker",
    ]
    sheet_from_rows("Factory_Order", money_rows, cols)

    # Projection for all demand SKUs (including non money-makers) for transparency
    proj = sorted(rows, key=lambda r: (-r["order_pcs"], -r["sold_hist"]))
    sheet_from_rows("Full_Projection", proj, cols)

    # Top-up suggestions to fill container
    tu = wb.create_sheet("TopUp_To_Fill_Container")
    tu["A1"] = "Remaining cube after core order (45'HQ)"
    tu["B1"] = round(max(0, CBM_45HQ - cbm_known), 2)
    tu["A2"] = "Remaining cube after core order (40')"
    tu["B2"] = round(max(0, CBM_40 - cbm_known), 2)
    tu["A4"] = "Suggested top-ups (high margin, not in core order or deepen top sellers)"
    tu["A4"].font = Font(bold=True)

    # Deepen top 10 money-maker sellers by +2 months
    deepen = []
    top_sellers = sorted(
        [r for r in rows if r["money_maker"] and r["sold_hist"] > 0],
        key=lambda r: -r["sold_hist"],
    )[:12]
    for r in top_sellers:
        extra = ceil_pack(r["monthly"] * 2, r["pack"])
        if extra <= 0:
            continue
        c = data["cost"].get((r["code"], r["size"])) or {}
        cbm_pc = c.get("CBM_per_Pc") or 0
        deepen.append(
            {
                "action": "Deepen +2 mo",
                "code": r["code"],
                "tommur": r["tommur"],
                "size": r["size"],
                "extra_pcs": extra,
                "cartons": extra // r["pack"],
                "cbm": round(cbm_pc * extra, 4) if cbm_pc else None,
                "margin_pct": r["margin_pct"],
                "note": "Extra months of proven seller",
            }
        )

    # High-margin FOB SKUs with no sales
    opp = []
    for (code, size), mg in data["margins"].items():
        if not mg.get("margin_pct") or mg["margin_pct"] < 50:
            continue
        if (code, size) in data["sales"] and data["sales"][(code, size)]["qty"] > 0:
            continue
        c = data["cost"].get((code, size)) or {}
        m = data["meta"].get((code, size)) or {}
        pack = int(c.get("Pcs_per_Carton") or m.get("pack") or 1) or 1
        # suggest 1 carton trial
        opp.append(
            {
                "action": "Trial 1 ctn",
                "code": code,
                "tommur": m.get("tommur") or c.get("Tommur_Code") or "",
                "size": size,
                "extra_pcs": pack,
                "cartons": 1,
                "cbm": round((c.get("CBM_per_Pc") or 0) * pack, 4) if c.get("CBM_per_Pc") else None,
                "margin_pct": mg["margin_pct"],
                "note": f"No sales yet; online margin {mg['margin_pct']:.0f}%",
            }
        )
    opp = sorted(opp, key=lambda x: -(x["margin_pct"] or 0))[:25]

    headers = ["Action", "Code", "Tommur", "Size", "Extra_Pcs", "Cartons", "CBM", "Margin_%", "Note"]
    for c, h in enumerate(headers, 1):
        cell = tu.cell(6, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
    rr = 7
    for item in deepen + opp:
        vals = [
            item["action"],
            item["code"],
            item["tommur"],
            item["size"],
            item["extra_pcs"],
            item["cartons"],
            item["cbm"],
            item["margin_pct"],
            item["note"],
        ]
        for c, v in enumerate(vals, 1):
            tu.cell(rr, c, v).border = thin
        rr += 1
    for i, w in enumerate([14, 16, 18, 12, 10, 10, 10, 10, 40], 1):
        tu.column_dimensions[get_column_letter(i)].width = w

    # Assumptions
    notes = wb.create_sheet("Assumptions")
    for i, t in enumerate(
        [
            "Live sales/inventory/backorders pulled from production D1 (read-only one-shot export; export worker deleted after).",
            "Monthly demand = historical sold qty ÷ span months (~4.34). Dominated by NJPD orders; Aug APBS-000005 is ~46k pcs.",
            "Containers 3 & 4 from data/inbound-containers.json (ETA 2026-09-15) applied as inbound before calculating reorder.",
            "Open backorder = order_items.quantity − qty_shipped on non-cancelled orders.",
            "Money-maker filter: competitive margin ≥30%, or PVC DWV fitting (family peers are 70–90% vs online).",
            "Pipe FOB/landed/online are per foot; fittings per piece. Order quantities are pieces (or feet for pipe).",
            "CBM from Tommur cost tracker carton dims; SKUs without dims contribute pcs but not CBM fill %.",
            "Factory CSV is carton-rounded and ready to send to Tommur.",
        ],
        1,
    ):
        notes.cell(i, 1, t)
        notes.column_dimensions["A"].width = 120

    for path in (OUT_XLSX, ART_XLSX):
        try:
            wb.save(path)
        except OSError as e:
            print("WARN", path, e)

    print(f"Money-maker order lines={len(money_rows)} pcs={pcs} cartons={cartons}")
    print(f"CBM known={cbm_known:.2f}  fill40={fill40:.1f}%  fill45={fill45:.1f}%")
    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_CSV}")
    return money_rows, cbm_known, fill40, fill45


def main():
    data = load_all()
    rows = build_rows(data)
    money_rows, cbm, f40, f45 = write_outputs(data, rows)
    print("\nTop order lines:")
    for r in money_rows[:20]:
        print(
            f"  {r['code']} {r['size']}: order={r['order_pcs']} ctn={r['cartons']} "
            f"mo={r['monthly']} avail_after={r['available_after']} m%={r['margin_pct']}"
        )


if __name__ == "__main__":
    main()
