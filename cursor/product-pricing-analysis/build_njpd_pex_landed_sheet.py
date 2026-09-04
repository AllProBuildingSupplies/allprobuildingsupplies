#!/usr/bin/env python3
"""Build one workbook: Tommur DDP vs APBS FOB-landed at $2800+$500 / 40',
NJPD/Gator Everflow PEX sells, and SKU match vs what NJPD actually buys."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows  # noqa: F401  (compat)

ROOT = Path(__file__).resolve().parent
ART = Path("/opt/cursor/artifacts")
ART.mkdir(parents=True, exist_ok=True)

# --- Freight / duty assumptions ---
OCEAN_USD = 2800.0
DRAYAGE_USD = 500.0
FREIGHT_TOTAL = OCEAN_USD + DRAYAGE_USD  # 3300
CBM_40FT = 67.7  # tracker: 40' internal
FREIGHT_PER_CBM = FREIGHT_TOTAL / CBM_40FT
MFN = 0.053
SEC301 = 0.25
FLIP = 0.125
DUTY_PCT = MFN + SEC301 + FLIP  # 0.428
TOMMUR_19K_OCEAN = 9500.0
TOMMUR_19K_DUTY = 7900.0
TOMMUR_19K_DELIV = 1600.0

# --- NJPD / Gator documents ---
# Order ack 11234952: Everflow 20ft straight PEX-B sticks (unit = stick)
# Quote 11237145: Everflow poly-alloy F2159 fittings + copper F1807 crimp rings

NJPD_PIPE = [
    # gator, desc, size, color, stick_price, sticks, tommur_apbs
    ("PFW-B120", '1" BLUE PEX TUBING (20 FT STRAIGHT)', '1"', "Blue", 11.6250, 150, "PEX-B PIPE"),
    ("PFW-R120", '1" RED PEX TUBING (20 FT STRAIGHT)', '1"', "Red", 11.6250, 100, "PEX-B PIPE"),
    ("PFW-B3420", '3/4" BLUE PEX TUBING (20 FT STRAIGHT)', '3/4"', "Blue", 6.7000, 250, "PEX-B PIPE"),
    ("PFW-R3420", '3/4" RED PEX TUBING (20 FT STRAIGHT)', '3/4"', "Red", 6.7000, 250, "PEX-B PIPE"),
    ("PFW-B1220", '1/2" BLUE PEX TUBING (20 FT STRAIGHT)', '1/2"', "Blue", 3.5500, 300, "PEX-B PIPE"),
    ("PFW-R1220", '1/2" RED PEX TUBING (20 FT STRAIGHT)', '1/2"', "Red", 3.5500, 200, "PEX-B PIPE"),
]

# Fittings: gator, desc, size_norm, unit_price, qty, tommur_apbs or None, tommur_size or None
NJPD_FITTINGS = [
    ("PPLN0100", '1" POLY-ALLOY PEX ELBOW', '1"', 1.4607, 900, "PEX-ELBOW", '1"'),
    ("PPTE0100", '1" POLY-ALLOY PEX TEE', '1"', 2.0344, 200, "PEX-TEE", '1"'),
    ("PPRT1134", '1"x1"x3/4" POLY-ALLOY PEX REDUCING TEE', '1x1x3/4"', 1.5782, 200, "PEX-REDTEE", '1x1x3/4"'),
    ("PPRT1112", '1"x1"x1/2" POLY-ALLOY PEX REDUCING TEE', '1x1x1/2"', 1.3929, 300, None, None),
    ("PPRT1341", '1"x3/4"x1" POLY-ALLOY PEX REDUCING TEE', '1x3/4x1"', 2.0525, 200, None, None),
    ("PPRT1033", '1"x3/4"x3/4" POLY-ALLOY PEX REDUCING TEE', '1x3/4x3/4"', 1.6745, 300, None, None),
    ("PPCP0100", '1" POLY-ALLOY PEX COUPLING', '1"', 0.7695, 400, "PEX-CPLNG", '1"'),
    ("PPRC1034", '1"x3/4" POLY-ALLOY PEX REDUCING COUPLING', '1x3/4"', 0.7138, 600, None, None),
    ("PPPL0100", '1" POLY-ALLOY PEX PLUG', '1"', 0.5903, 500, None, None),
    ("PPLN0034", '3/4" POLY-ALLOY PEX ELBOW', '3/4"', 0.6581, 1500, "PEX-ELBOW", '3/4"'),
    ("PPTE0034", '3/4" POLY-ALLOY PEX TEE', '3/4"', 0.9231, 600, "PEX-TEE", '3/4"'),
    ("PPRT3410", '3/4"x3/4"x1" POLY-ALLOY PEX REDUCING TEE', '3/4x3/4x1"', 1.4366, 150, None, None),
    ("PPRT3431", '3/4"x3/4"x1/2" POLY-ALLOY PEX REDUCING TEE', '3/4x3/4x1/2"', 0.7439, 1000, "PEX-REDTEE", '3/4x3/4x1/2"'),
    ("PPRT3413", '3/4"x1/2"x3/4" POLY-ALLOY PEX REDUCING TEE', '3/4x1/2x3/4"', 0.7891, 750, None, None),
    ("PPRT3411", '3/4"x1/2"x1/2" POLY-ALLOY PEX REDUCING TEE', '3/4x1/2x1/2"', 0.5963, 900, None, None),
    ("PPCP0034", '3/4" POLY-ALLOY PEX COUPLING', '3/4"', 0.3795, 1500, "PEX-CPLNG", '3/4"'),
    ("PPRC3412", '3/4"x1/2" POLY-ALLOY PEX REDUCING COUPLING', '3/4x1/2"', 0.3313, 600, "PEX-REDUCER", '3/4x1/2"'),
    ("PPPL0034", '3/4" POLY-ALLOY PEX PLUG', '3/4"', 0.3629, 1000, None, None),
    ("PPLN0012", '1/2" POLY-ALLOY PEX ELBOW', '1/2"', 0.3313, 3000, "PEX-ELBOW", '1/2"'),
    ("PPTE0012", '1/2" POLY-ALLOY PEX TEE', '1/2"', 0.4488, 400, "PEX-TEE", '1/2"'),
    ("PPRT1213", '1/2"x1/2"x3/4" POLY-ALLOY PEX REDUCING TEE', '1/2x1/2x3/4"', 0.6445, 300, None, None),
    ("PPCP0012", '1/2" POLY-ALLOY PEX COUPLING', '1/2"', 0.2304, 1000, "PEX-CPLNG", '1/2"'),
    ("PPPL0012", '1/2" POLY-ALLOY PEX PLUG', '1/2"', 0.2078, 2000, None, None),
    ("EPCR0012", '1/2" COPPER PEX CRIMP RING', '1/2"', 0.1529, 10000, None, None),
    ("EPCR0034", '3/4" COPPER PEX CRIMP RING', '3/4"', 0.2118, 6000, None, None),
    ("EPCR0100", '1" COPPER PEX CRIMP RING', '1"', 0.3765, 3500, None, None),
]


def nsize(s):
    return str(s or "").replace('"', "").replace(" ", "").lower()


def money(x):
    if x is None or x == "":
        return None
    try:
        return round(float(x), 4)
    except (TypeError, ValueError):
        return None


def pct(num, den):
    if num is None or den in (None, 0):
        return None
    return round(num / den * 100, 1)


def load_sources():
    analysis = {}
    with open(ROOT / "Tommur_Updated_DDP_Margin_Analysis.csv") as f:
        for r in csv.DictReader(f):
            analysis[(r["APBS_Code"].strip(), nsize(r["Size"]))] = r

    wb = openpyxl.load_workbook("/opt/cursor/artifacts/Factory_Order_PVC_PEX_45HQ.xlsx", data_only=True)
    ws = wb["Full_Pricing"]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else "" for h in rows[0]]
    fp = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        if d.get("APBS_Code"):
            fp.append(d)
    return analysis, fp


def landed(fob, cbm, qty):
    """FOB + 42.8% duty/tariff + freight/drayage allocated by CBM."""
    fob = money(fob)
    if fob is None:
        return None, None, None, None
    duty = round(fob * DUTY_PCT, 4)
    fr = None
    if cbm not in (None, "") and qty:
        try:
            fr = round((float(cbm) / float(qty)) * FREIGHT_PER_CBM, 4)
        except (TypeError, ValueError, ZeroDivisionError):
            fr = None
    landed_full = round(fob + duty + (fr or 0), 4)
    landed_fr_only = round(fob + (fr or 0), 4)  # what $3300-vs-$19k intuition often skips
    return duty, fr, landed_full, landed_fr_only


def build_rows(analysis, fp):
    # Index NJPD fittings by tommur key
    njpd_fit_by_tommur = {}
    for g, desc, sz, price, qty, apbs, tsz in NJPD_FITTINGS:
        if apbs:
            njpd_fit_by_tommur[(apbs, nsize(tsz))] = {
                "gator": g, "desc": desc, "size": sz, "price": price, "qty": qty
            }

    pipe_by_size = {}
    for g, desc, sz, color, stick, sticks, apbs in NJPD_PIPE:
        pipe_by_size.setdefault(nsize(sz), []).append(
            dict(gator=g, desc=desc, size=sz, color=color, stick=stick, sticks=sticks, apbs=apbs)
        )

    # Tommur PEX keys to DROP (NJPD does not buy these sizes/types)
    drop_pex = set()
    keep_pex = set()
    for d in fp:
        if d["Material"] != "PEX":
            continue
        k = (d["APBS_Code"], nsize(d["Size"]))
        code, sz = k
        if code == "PEX-B PIPE" and sz in ("1/2", "3/4", "1"):
            keep_pex.add(k)
        elif code == "PEX-ELBOW" and sz in ("1/2", "3/4"):  # 1" not on Tommur list
            keep_pex.add(k)
        elif code in ("PEX-CPLNG", "PEX-TEE") and sz in ("1/2", "3/4", "1"):
            keep_pex.add(k)
        elif code == "PEX-REDTEE" and sz in ("1x1x3/4", "3/4x3/4x1/2"):
            keep_pex.add(k)
        elif code == "PEX-REDUCER" and sz == "3/4x1/2":
            keep_pex.add(k)
        else:
            drop_pex.add(k)

    rows = []

    # PVC — keep factory-order lines, All Pro sell
    for d in fp:
        if d["Material"] != "PVC":
            continue
        k = (d["APBS_Code"], nsize(d["Size"]))
        a = analysis.get(k, {})
        fob = money(a.get("FOB_USD") or d.get("FOB_USD"))
        ddp = money(a.get("DDP_USD"))
        if ddp is None and d.get("Material") == "PVC":
            ddp = money(d.get("DDP_USD"))
        cur = money(a.get("CURRENT_DDP_USD"))
        sell = money(d.get("AllPro_Sell"))
        qty = d.get("Order_Pcs") or 0
        cbm = d.get("CBM")
        duty, fr, land, land_ex = landed(fob, cbm, qty)
        tommur_cost = ddp if ddp is not None else cur
        rows.append(line(
            d, a, fob, ddp, cur, sell, qty, cbm, duty, fr, land, land_ex, tommur_cost,
            sell_source="All Pro website",
            njpd=None,
            action="KEEP — PVC (All Pro sell)",
            match="PVC — not on NJPD PEX docs",
        ))

    # PEX keep — overlay NJPD sell + NJPD qty
    for d in fp:
        if d["Material"] != "PEX":
            continue
        k = (d["APBS_Code"], nsize(d["Size"]))
        a = analysis.get(k, {})
        fob = money(a.get("FOB_USD") or d.get("FOB_USD"))
        ddp = money(a.get("DDP_USD"))  # blank on many new PEX fittings — do not reuse old factory DDP
        cur = money(a.get("CURRENT_DDP_USD"))
        cbm = d.get("CBM")
        tommur_cost = ddp if ddp is not None else cur

        if k in drop_pex:
            qty = d.get("Order_Pcs") or 0
            duty, fr, land, land_ex = landed(fob, cbm, qty)
            rows.append(line(
                d, a, fob, ddp, cur, None, qty, cbm, duty, fr, land, land_ex, tommur_cost,
                sell_source=None,
                njpd=None,
                action="DROP — NJPD does not buy this size/type",
                match="No match on Gator/NJPD PEX list",
            ))
            continue

        if d["APBS_Code"] == "PEX-B PIPE":
            colors = pipe_by_size.get(nsize(d["Size"]), [])
            ft = sum(c["sticks"] * 20 for c in colors)
            stick = colors[0]["stick"] if colors else None
            sell_ft = round(stick / 20, 4) if stick else None
            gators = ", ".join(c["gator"] for c in colors)
            color_note = " + ".join(f"{c['color']} {c['sticks']} sticks" for c in colors)
            duty, fr, land, land_ex = landed(fob, cbm, ft or d.get("Order_Pcs"))
            # scale CBM to new qty vs old factory qty
            old_qty = d.get("Order_Pcs") or 0
            new_cbm = None
            if cbm and old_qty and ft:
                new_cbm = round(float(cbm) * (ft / float(old_qty)), 4)
                duty, fr, land, land_ex = landed(fob, new_cbm, ft)
            rec = line(
                d, a, fob, ddp, cur, sell_ft, ft, new_cbm or cbm, duty, fr, land, land_ex, tommur_cost,
                sell_source="NJPD/Gator (per ft = 20ft stick ÷ 20)",
                njpd={
                    "Gator_Code": gators,
                    "NJPD_Desc": colors[0]["desc"] if colors else "",
                    "NJPD_Stick_USD": stick,
                    "NJPD_Color_Split": color_note,
                    "NJPD_Sticks": sum(c["sticks"] for c in colors),
                },
                action="KEEP — match NJPD 20ft red/blue sticks (order in feet)",
                match="MATCH pipe size — ASK factory for 20ft STRAIGHT red + blue (not coil/natural)",
            )
            rec["Unit"] = "ft"
            rows.append(rec)
            continue

        nj = njpd_fit_by_tommur.get(k)
        if not nj:
            qty = d.get("Order_Pcs") or 0
            duty, fr, land, land_ex = landed(fob, cbm, qty)
            rows.append(line(
                d, a, fob, ddp, cur, None, qty, cbm, duty, fr, land, land_ex, tommur_cost,
                sell_source=None,
                njpd=None,
                action="DROP — not on NJPD quote",
                match="Tommur has it; NJPD does not buy",
            ))
            continue

        qty = nj["qty"]
        old_qty = d.get("Order_Pcs") or 0
        new_cbm = None
        if cbm and old_qty and qty:
            new_cbm = round(float(cbm) * (qty / float(old_qty)), 4)
        duty, fr, land, land_ex = landed(fob, new_cbm or cbm, qty)
        rows.append(line(
            d, a, fob, ddp, cur, nj["price"], qty, new_cbm or cbm, duty, fr, land, land_ex, tommur_cost,
            sell_source="NJPD/Gator quote 11237145",
            njpd={
                "Gator_Code": nj["gator"],
                "NJPD_Desc": nj["desc"],
                "NJPD_Stick_USD": None,
                "NJPD_Color_Split": "",
                "NJPD_Sticks": None,
            },
            action="KEEP — exact NJPD fitting",
            match="MATCH poly-alloy crimp fitting (must be ASTM F2159)",
        ))

    # NJPD items Tommur does not have
    have = {(r["APBS_Code"], nsize(r["Size"])) for r in rows if r["Material"] == "PEX" and r["Action"].startswith("KEEP")}
    # pipe handled as one row per size
    for g, desc, sz, price, qty, apbs, tsz in NJPD_FITTINGS:
        if apbs and (apbs, nsize(tsz)) in have:
            continue
        if apbs == "PEX-ELBOW" and nsize(sz) == "1":
            # Tommur factory order omitted 1" elbow — still ask
            action = "ADD — NJPD buys 1\" elbow; not on current Tommur order list"
        elif g.startswith("EPCR"):
            action = "ADD / SOURCE ELSEWHERE — copper F1807 crimp rings (Tommur PEX list has none)"
        else:
            action = "ASK TOMMUR — NJPD buys this; not on Tommur quote"
        rows.append({
            "Action": action,
            "Match": "MISSING from Tommur list",
            "Tommur_Code": "",
            "APBS_Code": apbs or g,
            "Gator_Code": g,
            "NJPD_Desc": desc,
            "Description": desc,
            "Size": sz if sz.endswith('"') else sz,
            "Material": "PEX" if not g.startswith("EPCR") else "COPPER",
            "Color": "",
            "Unit": "pc",
            "Order_Qty": qty,
            "CBM": None,
            "FOB_USD": None,
            "Duty_per_unit": None,
            "Freight_dray_per_unit": None,
            "Landed_APBS_USD": None,
            "Landed_FOB_plus_freight_only": None,
            "DDP_USD": None,
            "CURRENT_DDP_USD": None,
            "Tommur_cost_used": None,
            "Better_buy": "Need Tommur price",
            "Landed_vs_DDP_USD": None,
            "Sell_USD": price,
            "Sell_Source": "NJPD/Gator quote 11237145",
            "NJPD_Stick_USD": None,
            "Margin_on_DDP_pct": None,
            "Margin_on_Landed_pct": None,
            "Line_cost_DDP": None,
            "Line_cost_Landed": None,
            "Line_revenue": round(price * qty, 2),
            "Line_profit_DDP": None,
            "Line_profit_Landed": None,
            "Notes": "Not on Tommur PVC/PEX quote. Confirm F2159 poly-alloy (fittings) or F1807 copper ring.",
        })

    return rows, drop_pex, keep_pex


def line(d, a, fob, ddp, cur, sell, qty, cbm, duty, fr, land, land_ex, tommur_cost,
         sell_source, njpd, action, match):
    njpd = njpd or {}
    better = None
    delta = None
    if land is not None and tommur_cost is not None:
        delta = round(land - tommur_cost, 4)
        better = "FOB+landed cheaper" if land < tommur_cost else ("DDP cheaper" if land > tommur_cost else "tie")
    elif tommur_cost is not None:
        better = "DDP (no FOB to land)"
    elif land is not None:
        better = "FOB landed (no DDP)"

    m_ddp = pct((sell - tommur_cost) if sell and tommur_cost else None, sell)
    m_land = pct((sell - land) if sell and land else None, sell)

    def lx(cost):
        if cost is None or not qty:
            return None
        return round(cost * qty, 2)

    return {
        "Action": action,
        "Match": match,
        "Tommur_Code": d.get("Tommur_Code") or a.get("Tommur_Code"),
        "APBS_Code": d["APBS_Code"],
        "Gator_Code": njpd.get("Gator_Code", ""),
        "NJPD_Desc": njpd.get("NJPD_Desc", ""),
        "Description": d.get("Description"),
        "Size": d.get("Size"),
        "Material": d.get("Material"),
        "Color": njpd.get("NJPD_Color_Split", ""),
        "Unit": d.get("Unit"),
        "Order_Qty": qty,
        "CBM": cbm,
        "FOB_USD": fob,
        "Duty_per_unit": duty,
        "Freight_dray_per_unit": fr,
        "Landed_APBS_USD": land,
        "Landed_FOB_plus_freight_only": land_ex,
        "DDP_USD": ddp,
        "CURRENT_DDP_USD": cur,
        "Tommur_cost_used": tommur_cost,
        "Better_buy": better,
        "Landed_vs_DDP_USD": delta,
        "Sell_USD": sell,
        "Sell_Source": sell_source,
        "NJPD_Stick_USD": njpd.get("NJPD_Stick_USD"),
        "Margin_on_DDP_pct": m_ddp,
        "Margin_on_Landed_pct": m_land,
        "Line_cost_DDP": lx(tommur_cost),
        "Line_cost_Landed": lx(land),
        "Line_revenue": lx(sell),
        "Line_profit_DDP": round((sell - tommur_cost) * qty, 2) if sell and tommur_cost and qty else None,
        "Line_profit_Landed": round((sell - land) * qty, 2) if sell and land and qty else None,
        "Notes": "",
    }


HEADER_FILL = PatternFill("solid", "FF1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Border(
    left=Side(style="thin", color="FFD0D0D0"),
    right=Side(style="thin", color="FFD0D0D0"),
    top=Side(style="thin", color="FFD0D0D0"),
    bottom=Side(style="thin", color="FFD0D0D0"),
)
RED = PatternFill("solid", "FFFFC7CE")
YELLOW = PatternFill("solid", "FFFFEB9C")
GREEN = PatternFill("solid", "FFC6EFCE")
GRAY = PatternFill("solid", "FFD9D9D9")
BLUE = PatternFill("solid", "FFDDEBF7")
ORANGE = PatternFill("solid", "FFFCE4D6")


def write_sheet(ws, headers, rows, freeze=True):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[1].height = 32
    for r_i, row in enumerate(rows, 2):
        action = str(row.get("Action") or "")
        for c, h in enumerate(headers, 1):
            val = row.get(h)
            cell = ws.cell(r_i, c, val)
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if h in ("Margin_on_DDP_pct", "Margin_on_Landed_pct") and isinstance(val, (int, float)):
                if val < 0:
                    cell.fill = RED
                elif val < 30:
                    cell.fill = YELLOW
                elif val >= 50:
                    cell.fill = GREEN
            if h == "Better_buy" and val == "DDP cheaper":
                cell.fill = GREEN
            if h == "Better_buy" and val == "FOB+landed cheaper":
                cell.fill = BLUE
            if h == "Action" and action.startswith("DROP"):
                cell.fill = GRAY
            if h == "Action" and action.startswith("ASK") or (h == "Action" and action.startswith("ADD")):
                cell.fill = ORANGE
        if action.startswith("DROP"):
            for c in range(1, len(headers) + 1):
                if ws.cell(r_i, c).fill.fgColor is None or ws.cell(r_i, c).fill.fgColor.rgb in ("00000000", None):
                    pass
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(36, len(h) + 3))
    if freeze:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        ws.freeze_panes = "A2"


def summarize(rows):
    keep = [r for r in rows if str(r["Action"]).startswith("KEEP")]
    pvc = [r for r in keep if r["Material"] == "PVC"]
    pex = [r for r in keep if r["Material"] == "PEX"]
    drop = [r for r in rows if str(r["Action"]).startswith("DROP")]
    missing = [r for r in rows if str(r["Action"]).startswith("ASK") or str(r["Action"]).startswith("ADD")]

    def roll(subset, cost_key, profit_key, margin_key):
        priced = [r for r in subset if r.get(cost_key) is not None and r.get("Line_revenue") is not None]
        cost = sum(r[cost_key] or 0 for r in priced)
        rev = sum(r["Line_revenue"] or 0 for r in priced)
        profit = sum(r[profit_key] or 0 for r in priced)
        ms = [r[margin_key] for r in priced if r.get(margin_key) is not None]
        ms_sorted = sorted(ms)
        med = ms_sorted[len(ms_sorted) // 2] if ms_sorted else None
        under = sum(1 for m in ms if m < 0)
        return dict(n=len(priced), cost=cost, rev=rev, profit=profit,
                    order_m=pct(profit, rev), median=med, min=min(ms) if ms else None,
                    underwater=under)

    ddp_pvc = roll(pvc, "Line_cost_DDP", "Line_profit_DDP", "Margin_on_DDP_pct")
    land_pvc = roll([r for r in pvc if r.get("FOB_USD")], "Line_cost_Landed", "Line_profit_Landed", "Margin_on_Landed_pct")
    ddp_pex = roll(pex, "Line_cost_DDP", "Line_profit_DDP", "Margin_on_DDP_pct")
    land_pex = roll([r for r in pex if r.get("FOB_USD")], "Line_cost_Landed", "Line_profit_Landed", "Margin_on_Landed_pct")

    # FOB merchandise + duty + freight for keep lines with FOB
    fob_keep = [r for r in keep if r.get("FOB_USD") and r.get("Order_Qty")]
    fob_merch = sum(r["FOB_USD"] * r["Order_Qty"] for r in fob_keep)
    duty_tot = fob_merch * DUTY_PCT
    cbm_keep = sum(float(r["CBM"]) for r in keep if r.get("CBM"))
    ddp_buy = [r for r in keep if r.get("Tommur_cost_used") and r.get("Order_Qty")]
    ddp_tot = sum(r["Tommur_cost_used"] * r["Order_Qty"] for r in ddp_buy)

    better_ddp = sum(1 for r in keep if r.get("Better_buy") == "DDP cheaper")
    better_fob = sum(1 for r in keep if r.get("Better_buy") == "FOB+landed cheaper")

    pex_under_ddp = [r for r in pex if r.get("Margin_on_DDP_pct") is not None and r["Margin_on_DDP_pct"] < 0]
    pex_under_land = [r for r in pex if r.get("Margin_on_Landed_pct") is not None and r["Margin_on_Landed_pct"] < 0]

    return {
        "keep": keep, "pvc": pvc, "pex": pex, "drop": drop, "missing": missing,
        "ddp_pvc": ddp_pvc, "land_pvc": land_pvc, "ddp_pex": ddp_pex, "land_pex": land_pex,
        "fob_merch": fob_merch, "duty_tot": duty_tot, "cbm_keep": cbm_keep,
        "ddp_tot": ddp_tot, "better_ddp": better_ddp, "better_fob": better_fob,
        "pex_under_ddp": pex_under_ddp, "pex_under_land": pex_under_land,
        "fob_keep_n": len(fob_keep),
    }


def write_summary(ws, s):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 88
    title = Font(bold=True, size=14, color="1F4E79")
    h = Font(bold=True, size=12, color="1F4E79")
    bold = Font(bold=True)

    lines = [
        ("Tommur DDP vs APBS FOB-landed  |  NJPD Everflow PEX match", title),
        ("", None),
        ("FREIGHT — $19k is NOT replaced by $3,300", h),
        ("Tommur 40HQ stack", None),
        ("  Sea freight", f"${TOMMUR_19K_OCEAN:,.0f}"),
        ("  Customs duties", f"${TOMMUR_19K_DUTY:,.0f}  ← still due if you buy FOB"),
        ("  Delivery", f"${TOMMUR_19K_DELIV:,.0f}"),
        ("  Total Tommur quoted", "$19,000 / 40HQ"),
        ("Your new quotes", None),
        ("  Ocean 40'", f"${OCEAN_USD:,.0f}"),
        ("  Drayage", f"${DRAYAGE_USD:,.0f}"),
        ("  Freight+drayage", f"${FREIGHT_TOTAL:,.0f}  (replaces $9,500+$1,600 only)"),
        ("  Duties on FOB (est. 42.8% = 5.3% MFN + 25% 301 + 12.5% FLIP)", "see merchandise calc below"),
        ("40' internal CBM used", f"{CBM_40FT} → freight ${FREIGHT_PER_CBM:.2f}/CBM"),
        ("", None),
        ("CONTAINER MATH ON KEEP LINES (PVC + NJPD-matched PEX with FOB)", h),
        ("FOB merchandise", f"${s['fob_merch']:,.0f}  ({s['fob_keep_n']} SKUs)"),
        ("Est. duty/tariff 42.8%", f"${s['duty_tot']:,.0f}"),
        ("Freight+drayage (if box is full 67.7 CBM)", f"${FREIGHT_TOTAL:,.0f}"),
        ("Keep-list CBM", f"{s['cbm_keep']:.1f} CBM  ({'fits one 40' if s['cbm_keep']<=CBM_40FT else 'OVER one 40 — need 40HQ/45HQ or 2 boxes'})"),
        ("DIY landed (FOB+duty+$3,300 allocated)", f"${s['fob_merch']+s['duty_tot']+FREIGHT_TOTAL:,.0f}  (duty dominates; $3,300 is small)"),
        ("Tommur DDP buy (same SKUs)", f"${s['ddp_tot']:,.0f}"),
        ("Winner", "DDP still cheaper on most lines because Tommur DDP is only ~5–15% over FOB while US duties are ~43%"),
        ("SKU count DDP cheaper / FOB-landed cheaper", f"{s['better_ddp']} / {s['better_fob']}"),
        ("", None),
        ("PVC @ All Pro website sells (KEEP)", h),
        ("n / cost DDP / revenue / profit / order margin / SKU median",
         f"{s['ddp_pvc']['n']}  |  ${s['ddp_pvc']['cost']:,.0f}  |  ${s['ddp_pvc']['rev']:,.0f}  |  ${s['ddp_pvc']['profit']:,.0f}  |  {s['ddp_pvc']['order_m']}%  |  {s['ddp_pvc']['median']}%"),
        ("Same on FOB-landed (where FOB exists)",
         f"{s['land_pvc']['n']}  |  ${s['land_pvc']['cost']:,.0f}  |  ${s['land_pvc']['rev']:,.0f}  |  ${s['land_pvc']['profit']:,.0f}  |  {s['land_pvc']['order_m']}%  |  {s['land_pvc']['median']}%"),
        ("", None),
        ("PEX @ NJPD/Gator prices (KEEP matches only)", h),
        ("n / cost DDP / revenue / profit / order margin / SKU median",
         f"{s['ddp_pex']['n']}  |  ${s['ddp_pex']['cost']:,.0f}  |  ${s['ddp_pex']['rev']:,.0f}  |  ${s['ddp_pex']['profit']:,.0f}  |  {s['ddp_pex']['order_m']}%  |  {s['ddp_pex']['median']}%"),
        ("Same on FOB-landed",
         f"{s['land_pex']['n']}  |  ${s['land_pex']['cost']:,.0f}  |  ${s['land_pex']['rev']:,.0f}  |  ${s['land_pex']['profit']:,.0f}  |  {s['land_pex']['order_m']}%  |  {s['land_pex']['median']}%"),
        ("PEX underwater on Tommur DDP", str(len(s["pex_under_ddp"])) + " SKUs — fittings cannot match Gator street at current Tommur DDP"),
        ("PEX underwater on FOB-landed", str(len(s["pex_under_land"])) + " SKUs"),
        ("", None),
        ("NJPD PRODUCT FILTER", h),
        ("KEEP PEX", f"{len(s['pex'])} Tommur lines that match 1/2–1\" pipe + poly crimp fittings NJPD buys"),
        ("DROP PEX", f"{len([r for r in s['drop'] if r['Material']=='PEX'])} lines (1-1/4\"+ pipe/elbows, adapters, extra reducers)"),
        ("ASK / ADD", f"{len(s['missing'])} NJPD items Tommur did not quote (plugs, extra reducing tees, 1\" elbow, 1x3/4 coupler, copper rings)"),
        ("", None),
        ("EVERFLOW SPECS TO SEND TOMMUR", h),
        ("Pipe brand NJPD uses", "Everflow (Gator PFW-*) — listed as PEX-B, 20 ft STRAIGHT sticks, red & blue"),
        ("Pipe standards", "ASTM F876 / F877, CSA B137.5, NSF/ANSI 14 & 61 (NSF-pw), SDR-9"),
        ("Pipe ratings (typical PEX-B SDR-9)", "160 psi @ 73°F, 100 psi @ 180°F; chlorine resistance ASTM F2023"),
        ("Pipe form", "20 ft straight (NOT coil). Hot=red, cold=blue. Non-barrier potable."),
        ("Pipe sizes NJPD buys", '1/2", 3/4", 1" ONLY — drop 1-1/4 / 1-1/2 / 2"'),
        ("Fittings brand", "Everflow Poly-Alloy (PP* codes) — PPSU/Acudel plastic insert"),
        ("Fittings standard", "ASTM F2159 (plastic insert + copper crimp ring) + F877, CSA B137.5, NSF-pw"),
        ("NOT F1960", "These are CRIMP fittings, not PEX-A expansion (F1960 / Uponor-style)"),
        ("Crimp rings", "Copper ASTM F1807 rings (Gator EPCR*) — separate from poly fittings"),
        ("Fitting pressure", "100 psi @ 180°F typical for F2159 poly-alloy"),
        ("", None),
        ("WHAT TO TELL THE FACTORY", h),
        ("1", "PEX-B silane, SDR-9, F876/F877, NSF-pw, 20ft straight red+blue 1/2 3/4 1 only"),
        ("2", "Fittings must be ASTM F2159 poly (PPSU) crimp, copper-ring; not brass F1807 body, not F1960"),
        ("3", "Quote the MISSING NJPD SKUs (plugs, reducing tees listed on Match sheet, 1\" elbow, 1x3/4 coupler)"),
        ("4", "DDP on 1/2\" and 3/4\" elbows ($0.49 / $0.83) is ABOVE what NJPD pays Gator ($0.33 / $0.66) — need a fittings DDP cut or those SKUs lose money"),
        ("5", "Pipe DDP still works at NJPD stick prices (~44–47% margin)"),
    ]
    r = 1
    for a, b in lines:
        if isinstance(b, Font):
            ws.cell(r, 1, a).font = b
        elif b is None and a:
            ws.cell(r, 1, a).font = h if a == a else bold
            if a:
                ws.cell(r, 1).font = h
        else:
            ws.cell(r, 1, a).font = bold if a and not a.startswith(" ") else Font()
            ws.cell(r, 2, b)
        r += 1
    ws.freeze_panes = "A3"


def main():
    analysis, fp = load_sources()
    rows, drop_pex, keep_pex = build_rows(analysis, fp)
    s = summarize(rows)

    headers = [
        "Action", "Match", "Tommur_Code", "APBS_Code", "Gator_Code", "NJPD_Desc",
        "Description", "Size", "Material", "Color", "Unit", "Order_Qty", "CBM",
        "FOB_USD", "Duty_per_unit", "Freight_dray_per_unit", "Landed_APBS_USD",
        "Landed_FOB_plus_freight_only", "DDP_USD", "CURRENT_DDP_USD", "Tommur_cost_used",
        "Better_buy", "Landed_vs_DDP_USD", "Sell_USD", "Sell_Source", "NJPD_Stick_USD",
        "Margin_on_DDP_pct", "Margin_on_Landed_pct",
        "Line_cost_DDP", "Line_cost_Landed", "Line_revenue",
        "Line_profit_DDP", "Line_profit_Landed", "Notes",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"
    write_summary(ws, s)

    # Combined: KEEP first, then ASK, then DROP
    def rank(r):
        a = r["Action"]
        if a.startswith("KEEP") and r["Material"] == "PEX":
            return 0
        if a.startswith("KEEP"):
            return 1
        if a.startswith("ASK") or a.startswith("ADD"):
            return 2
        return 3
    ordered = sorted(rows, key=lambda r: (rank(r), r["Material"], r["APBS_Code"], str(r["Size"])))

    ws2 = wb.create_sheet("Combined_List")
    write_sheet(ws2, headers, ordered)

    keep_only = [r for r in ordered if str(r["Action"]).startswith("KEEP")]
    ws3 = wb.create_sheet("Order_KEEP_only")
    write_sheet(ws3, headers, keep_only)

    pex_rows = [r for r in ordered if r["Material"] in ("PEX", "COPPER")]
    ws4 = wb.create_sheet("NJPD_PEX_Match")
    write_sheet(ws4, headers, pex_rows)

    # Specs sheet
    ws5 = wb.create_sheet("Everflow_Specs")
    specs = [
        ["Field", "NJPD / Everflow (from Gator packing slip + quote)", "What Tommur must match"],
        ["Brand on slip", "Everflow (Gator PFW-* pipe, PP* fittings, EPCR* rings)", "Equivalent OEM OK if standards match"],
        ["Pipe type", "PEX-B (Gator listing: “PEX-B Tubing – Potable Water”)", "PEX-b silane crosslink — NOT PEX-A (Engel), NOT PEX-C"],
        ["Pipe form", "20 ft STRAIGHT sticks", "Straight 20 ft — NJPD is not buying coils on this PO"],
        ["Pipe color", "Red (hot) and Blue (cold) — they order both", "Print/stripe red + blue; natural-only will not match the job"],
        ["Pipe sizes used", '1/2", 3/4", 1" only', "Do not produce 1-1/4 / 1-1/2 / 2 for this customer"],
        ["Dimension ratio", "SDR-9 (CTS OD: 1/2=0.625\", 3/4=0.875\", 1=1.125\")", "SDR-9 CTS, same OD/ID as ASTM F876"],
        ["Pipe ASTM", "F876 (tubing) + F877 (hot/cold water systems)", "Mark pipe F876/F877"],
        ["Pipe other listings", "CSA B137.5, NSF/ANSI 14 & 61 (NSF-pw), typically ASTM F2023 chlorine", "NSF-pw / UPC or IAPMO listing required for NJ potable"],
        ["Pressure (typical SDR-9)", "160 psi @ 73°F; 100 psi @ 180°F", "Same rating"],
        ["Barrier", "Non-barrier potable (not EVOH oxygen-barrier heat pipe)", "Potable non-barrier"],
        ["Fitting material", "Poly-alloy / PPSU (Acudel 22000) injection molded", "Plastic insert F2159 — not brass body"],
        ["Fitting ASTM", "F2159 (plastic insert + copper crimp ring) + F877", "Print F2159 on the fitting"],
        ["Fitting listings", "NSF-pw, CSA B137.5, lead-free", "Same"],
        ["Install method", "Copper crimp ring, full-circle crimp tool", "NOT expansion (F1960), NOT press/sharkbite"],
        ["Crimp rings", "Copper ASTM F1807 rings (Gator EPCR 1/2, 3/4, 1)", "Rings are metal; can be third-party if F1807 SDR-9"],
        ["Working (fittings)", "100 psi @ 180°F typical", "Same"],
        ["Sizes / configs NJPD buys", "See NJPD_PEX_Match sheet — many reducing-tee orientations", "Quote every missing tee/plug/coupler"],
        ["Gator docs", "Order ack 11234952 (pipe) + Quote 11237145 (fittings/rings)", ""],
    ]
    for i, row in enumerate(specs, 1):
        for c, v in enumerate(row, 1):
            cell = ws5.cell(i, c, v)
            cell.border = THIN
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if i == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
        ws5.row_dimensions[i].height = 36 if i > 1 else 22
    ws5.column_dimensions["A"].width = 24
    ws5.column_dimensions["B"].width = 62
    ws5.column_dimensions["C"].width = 52

    ws6 = wb.create_sheet("Assumptions")
    assumptions = [
        ["Item", "Value", "Note"],
        ["Ocean freight / 40'", OCEAN_USD, "User quote"],
        ["Drayage", DRAYAGE_USD, "User quote"],
        ["Freight+drayage total", FREIGHT_TOTAL, "Allocated by CBM across the box"],
        ["40' CBM", CBM_40FT, "Same as cost-tracker (not 40HQ ~76 CBM)"],
        ["Freight $/CBM", round(FREIGHT_PER_CBM, 4), "3300 / 67.7"],
        ["MFN duty", MFN, "HTS 3917 plastics pipe/fittings ~5.3%"],
        ["Section 301", SEC301, "China-origin add-on — confirm with broker"],
        ["FLIP 301", FLIP, "Additional ~12.5% from Jul 2026 — confirm with broker"],
        ["Duty+tariff used", DUTY_PCT, "If FLIP does not apply, landed falls ~12.5 pts"],
        ["Landed_APBS", "FOB × 1.428 + CBM_share × $48.74", "True door-ish cost if you self-enter"],
        ["Landed_FOB_plus_freight_only", "FOB + CBM freight (NO duty)", "Shown so $3,300 vs $19k is not confused with duty"],
        ["Tommur_cost_used", "DDP_USD if filled, else CURRENT_DDP", "Blank DDP on many new PEX fittings"],
        ["PEX sell", "Gator unit price to NJPD", "Pipe: stick/20 = per ft"],
        ["PVC sell", "All Pro website", "Unchanged"],
        ["PEX order qty", "NJPD ack/quote qty", "Pipe converted sticks→feet"],
        ["PVC order qty", "Existing 45HQ factory order", "Unchanged"],
    ]
    for i, row in enumerate(assumptions, 1):
        for c, v in enumerate(row, 1):
            cell = ws6.cell(i, c, v)
            cell.border = THIN
            if i == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
    for col, w in zip("ABC", (36, 28, 62)):
        ws6.column_dimensions[col].width = w

    xlsx = ROOT / "Tommur_NJPD_PEX_Landed_vs_DDP.xlsx"
    csv_path = ROOT / "Tommur_NJPD_PEX_Landed_vs_DDP.csv"
    wb.save(xlsx)
    with open(csv_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        w.writeheader()
        w.writerows(ordered)

    for p in (xlsx, csv_path):
        dest = ART / p.name
        dest.write_bytes(p.read_bytes())

    # Markdown for repo
    md = []
    md.append("# Tommur DDP vs APBS FOB-landed ($2,800 + $500) — NJPD Everflow PEX\n")
    md.append("## Freight: $19k is not $3,300\n")
    md.append("Tommur’s $19,000/40HQ = **$9,500 ocean + $7,900 duties + $1,600 delivery**.\n")
    md.append(f"Your quotes replace ocean+delivery only: **${OCEAN_USD:,.0f} + ${DRAYAGE_USD:,.0f} = ${FREIGHT_TOTAL:,.0f}**.\n")
    md.append("**Duties still apply on FOB** (~42.8% = 5.3% MFN + 25% 301 + 12.5% FLIP, confirm with broker).\n")
    md.append(f"- FOB merchandise (KEEP lines with FOB): **${s['fob_merch']:,.0f}**\n")
    md.append(f"- Est. duty: **${s['duty_tot']:,.0f}**\n")
    md.append(f"- Freight+drayage: **${FREIGHT_TOTAL:,.0f}**\n")
    md.append(f"- DIY landed: **${s['fob_merch']+s['duty_tot']+FREIGHT_TOTAL:,.0f}** vs Tommur DDP **${s['ddp_tot']:,.0f}**\n")
    md.append(f"- SKUs DDP cheaper: **{s['better_ddp']}** · FOB-landed cheaper: **{s['better_fob']}**\n")
    md.append("Tommur DDP is still only ~5–15% over FOB, so **DDP still wins** once real US duties are on the FOB path.\n")
    md.append("\n## NJPD PEX filter\n")
    md.append("NJPD buys Everflow **PEX-B 20 ft straight red/blue** in **1/2, 3/4, 1\" only**, plus **poly-alloy F2159 crimp fittings** and **copper F1807 rings**. Drop larger pipe/elbows and adapters.\n")
    md.append(f"- KEEP PEX matches: {len(s['pex'])}\n- DROP: {len([r for r in s['drop'] if r['Material']=='PEX'])}\n- ASK/ADD (Tommur didn’t quote): {len(s['missing'])}\n")
    md.append("\n## Margins at NJPD/Gator sells (PEX) vs All Pro (PVC)\n")
    md.append(f"- PVC DDP: order margin **{s['ddp_pvc']['order_m']}%**, median **{s['ddp_pvc']['median']}%**\n")
    md.append(f"- PEX DDP @ Gator prices: order margin **{s['ddp_pex']['order_m']}%**, median **{s['ddp_pex']['median']}%**, underwater **{len(s['pex_under_ddp'])}** SKUs\n")
    md.append("Pipe still works (~45% on DDP). **1/2\" and 3/4\" elbows lose money** at Gator ($0.33 / $0.66) vs Tommur DDP ($0.49 / $0.83).\n")
    md.append("\nFull workbook: `Tommur_NJPD_PEX_Landed_vs_DDP.xlsx`\n")
    (ROOT / "TOMMUR_NJPD_PEX_LANDED.md").write_text("".join(md))
    (ART / "TOMMUR_NJPD_PEX_LANDED.md").write_text("".join(md))

    print("KEEP pvc", len(s["pvc"]), "pex", len(s["pex"]), "drop", len(s["drop"]), "missing", len(s["missing"]))
    print("FOB merch", round(s["fob_merch"], 2), "duty", round(s["duty_tot"], 2), "ddp tot", round(s["ddp_tot"], 2))
    print("CBM keep", round(s["cbm_keep"], 2), "better ddp/fob", s["better_ddp"], s["better_fob"])
    print("PVC ddp", s["ddp_pvc"])
    print("PEX ddp", s["ddp_pex"])
    print("PEX land", s["land_pex"])
    print("PEX underwater DDP:")
    for r in s["pex_under_ddp"]:
        print(" ", r["APBS_Code"], r["Size"], "ddp", r["Tommur_cost_used"], "sell", r["Sell_USD"], "m", r["Margin_on_DDP_pct"])
    print("PEX underwater landed:")
    for r in s["pex_under_land"]:
        print(" ", r["APBS_Code"], r["Size"], "land", r["Landed_APBS_USD"], "sell", r["Sell_USD"], "m", r["Margin_on_Landed_pct"])
    print("PEX KEEP margins:")
    for r in s["pex"]:
        print(f"  {r['APBS_Code']:16} {r['Size']:12} FOB={r['FOB_USD']} DDP={r['DDP_USD']} LAND={r['Landed_APBS_USD']} sell={r['Sell_USD']} mDDP={r['Margin_on_DDP_pct']} mL={r['Margin_on_Landed_pct']} better={r['Better_buy']}")
    print("MISSING:")
    for r in s["missing"]:
        print(" ", r["Gator_Code"], r["Description"], r["Sell_USD"], r["Order_Qty"])
    print("wrote", xlsx)


if __name__ == "__main__":
    main()
