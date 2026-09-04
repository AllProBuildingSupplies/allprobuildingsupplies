#!/usr/bin/env python3
"""King Smart (Alibaba kingsmartplumbing) vs NJPD/Gator sell vs Tommur vs Zhenpeng."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

OUT = Path(__file__).resolve().parent
DUTY = 0.428
FREIGHT_FITTING = 0.03  # ~Tommur 1/2" elbow freight/pc at $3,300/40'

# NJPD/Gator sell (quote 11237145 / ack 11234952)
# King Smart published Made-in-China / indexed Alibaba ranges.
# Low = typical China FOB / large qty. High = small qty / USA warehouse.
ROWS = [
    # desc, gator, njpd_sell, tommur_ddp, tommur_fob, ks_low, ks_high, zp_fob, has_ks, notes
    (
        '1/2" poly-alloy elbow',
        "PPLN0012",
        0.3313,
        0.49,
        0.2933,
        None,
        None,
        0.10,
        "ASK — catalog has F2159 elbows; 1/2\" elbow not priced publicly",
        "Zhenpeng B121290 published $0.10/pc MOQ 3000. King Smart lists 3/4\" F2159 elbow only.",
    ),
    (
        '3/4" poly-alloy elbow',
        "PPLN0034",
        0.6581,
        0.83,
        0.40,
        0.51,
        0.83,
        None,
        "YES — 3/4 in F2159/Pex B Elbow, MOQ 30",
        "King Smart list $0.51–$0.83. High end = USA-stock / small qty.",
    ),
    (
        '1" poly-alloy elbow',
        "PPLN0100",
        1.46,
        None,
        None,
        None,
        None,
        None,
        "ASK — F2159 elbows in catalog; 1\" not priced",
        "Not on Tommur quote either.",
    ),
    (
        '1/2" poly-alloy coupling',
        "PPCP0012",
        0.2304,
        0.25,
        0.1576,
        None,
        None,
        None,
        "ASK — F2159 couplings exist; 1/2\" not priced",
        "",
    ),
    (
        '3/4" poly-alloy coupling',
        "PPCP0034",
        0.3795,
        0.40,
        0.2406,
        None,
        None,
        None,
        "ASK",
        "",
    ),
    (
        '1" poly-alloy coupling',
        "PPCP0100",
        0.7695,
        0.81,
        0.509,
        None,
        None,
        None,
        "ASK",
        "",
    ),
    (
        '1/2" poly-alloy tee',
        "PPTE0012",
        0.4488,
        0.47,
        0.2955,
        0.38,
        0.49,
        None,
        "YES — indexed Alibaba F2159 1/2\" tee $0.38–$0.49, MOQ 120",
        "Accio/Alibaba index of King Smart listing.",
    ),
    (
        '3/4" poly-alloy tee',
        "PPTE0034",
        0.9231,
        0.84,
        0.5254,
        None,
        None,
        None,
        "ASK",
        "",
    ),
    (
        '1" poly-alloy tee',
        "PPTE0100",
        2.0344,
        1.45,
        0.906,
        None,
        None,
        None,
        "ASK",
        "",
    ),
    (
        '3/4"x1/2"x3/4" reducing tee',
        "PPRT3413",
        0.79,
        None,
        None,
        0.65,
        1.04,
        None,
        "YES — 3/4\"*1/2\"*3/4\" F2159/Pex B Reducing Tee PPSU, MOQ 20",
        "Orientation must match Gator side×side×middle. Confirm in RFQ.",
    ),
    (
        '1/2"x3/4"x1/2" reducing tee',
        "PPRT1213",
        0.64,
        None,
        None,
        0.52,
        0.83,
        None,
        "YES — F2159 1/2\"*3/4\"*1/2\" reducing tee $0.52–$0.83, MOQ 30",
        "Confirm orientation vs NJPD PPRT1213.",
    ),
    (
        '1" poly-alloy plug',
        "PPPL0100",
        0.59,
        None,
        None,
        0.47,
        0.68,
        None,
        "YES — 1 Inch F2159/Pex B Plug Lead Free PPSU, MOQ 50",
        "",
    ),
    (
        '1/2" PEX-B 20ft stick (per ft)',
        "PFW-B1220 / PFW-R1220",
        0.1775,
        0.10,
        0.0915,
        None,
        None,
        None,
        "NO — not in King Smart catalog",
        "They sell fittings, tools, valves, hoses, rings — not 20ft red/blue sticks.",
    ),
    (
        '3/4" PEX-B 20ft stick (per ft)',
        "PFW-B3420 / PFW-R3420",
        0.335,
        0.18,
        0.1717,
        None,
        None,
        None,
        "NO",
        "",
    ),
    (
        '1" PEX-B 20ft stick (per ft)',
        "PFW-B120 / PFW-R120",
        0.5813,
        0.31,
        0.2941,
        None,
        None,
        None,
        "NO",
        "",
    ),
]


def landed(fob):
    if fob is None:
        return None
    return round(fob * (1 + DUTY) + FREIGHT_FITTING, 4)


def margin(sell, cost):
    if sell is None or cost is None or sell <= 0:
        return None
    return round((sell - cost) / sell * 100, 1)


def ceil_for(sell, target_pct):
    return round(sell * (1 - target_pct / 100), 4)


def verdict(sell, usa_cost, china_landed):
    need50 = ceil_for(sell, 50)
    # Prefer USA-stock cost if present (what user is looking at).
    cost = usa_cost if usa_cost is not None else china_landed
    if cost is None:
        return "NO PRICE — RFQ"
    m = margin(sell, cost)
    if m is None:
        return "NO PRICE"
    if m >= 50:
        return "HIT 50%+"
    if m >= 25:
        return "THIN — not 50%"
    if m >= 0:
        return "BARELY POSITIVE"
    return "UNDERWATER vs NJPD"


def csv_escape(v):
    s = "" if v is None else str(v)
    if any(c in s for c in ',\"\n'):
        return '"' + s.replace('"', '""') + '"'
    return s


def build_rows():
    out = []
    for (
        desc,
        gator,
        sell,
        tommur_ddp,
        tommur_fob,
        ks_low,
        ks_high,
        zp_fob,
        match,
        notes,
    ) in ROWS:
        ks_usa = ks_high  # treat high as USA warehouse / small qty delivered
        ks_fob = ks_low
        ks_landed = landed(ks_fob) if ks_fob is not None else None
        zp_landed = landed(zp_fob) if zp_fob is not None else None
        out.append(
            {
                "Item": desc,
                "Gator_Code": gator,
                "NJPD_Sell_USD": sell,
                "Cost_for_50pct_margin": ceil_for(sell, 50),
                "Cost_for_75pct_margin": ceil_for(sell, 75),
                "KingSmart_match": match,
                "KingSmart_FOB_low_USD": ks_fob,
                "KingSmart_USA_stock_high_USD": ks_usa,
                "KingSmart_FOB_landed_est": ks_landed,
                "Margin_on_USA_stock_pct": margin(sell, ks_usa),
                "Margin_on_KS_FOB_landed_pct": margin(sell, ks_landed),
                "Tommur_DDP_USD": tommur_ddp,
                "Margin_on_Tommur_DDP_pct": margin(sell, tommur_ddp),
                "Zhenpeng_FOB_USD": zp_fob,
                "Zhenpeng_landed_est": zp_landed,
                "Margin_on_Zhenpeng_landed_pct": margin(sell, zp_landed),
                "Verdict_at_USA_stock": verdict(sell, ks_usa, ks_landed),
                "Notes": notes,
            }
        )
    return out


HEADERS = [
    "Item",
    "Gator_Code",
    "NJPD_Sell_USD",
    "Cost_for_50pct_margin",
    "Cost_for_75pct_margin",
    "KingSmart_match",
    "KingSmart_FOB_low_USD",
    "KingSmart_USA_stock_high_USD",
    "KingSmart_FOB_landed_est",
    "Margin_on_USA_stock_pct",
    "Margin_on_KS_FOB_landed_pct",
    "Tommur_DDP_USD",
    "Margin_on_Tommur_DDP_pct",
    "Zhenpeng_FOB_USD",
    "Zhenpeng_landed_est",
    "Margin_on_Zhenpeng_landed_pct",
    "Verdict_at_USA_stock",
    "Notes",
]


def write_csv(rows, path):
    lines = [",".join(HEADERS)]
    for r in rows:
        lines.append(",".join(csv_escape(r[h]) for h in HEADERS))
    path.write_text("\n".join(lines) + "\n")


def write_xlsx(rows, path):
    if Workbook is None:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "KingSmart_vs_NJPD"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    red = PatternFill("solid", fgColor="F4CCCC")
    green = PatternFill("solid", fgColor="D9EAD3")
    yellow = PatternFill("solid", fgColor="FFF2CC")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    money_cols = {
        "NJPD_Sell_USD",
        "Cost_for_50pct_margin",
        "Cost_for_75pct_margin",
        "KingSmart_FOB_low_USD",
        "KingSmart_USA_stock_high_USD",
        "KingSmart_FOB_landed_est",
        "Tommur_DDP_USD",
        "Zhenpeng_FOB_USD",
        "Zhenpeng_landed_est",
    }
    pct_cols = {
        "Margin_on_USA_stock_pct",
        "Margin_on_KS_FOB_landed_pct",
        "Margin_on_Tommur_DDP_pct",
        "Margin_on_Zhenpeng_landed_pct",
    }

    for i, r in enumerate(rows, 2):
        for col, h in enumerate(HEADERS, 1):
            v = r[h]
            cell = ws.cell(i, col, v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            if h in money_cols and isinstance(v, (int, float)):
                cell.number_format = '"$"#,##0.0000'
            if h in pct_cols and isinstance(v, (int, float)):
                cell.number_format = '0.0"%"'
                if v < 0:
                    cell.fill = red
                elif v >= 50:
                    cell.fill = green
                elif v < 25:
                    cell.fill = yellow
            if h == "Verdict_at_USA_stock":
                if v == "UNDERWATER vs NJPD":
                    cell.fill = red
                elif v == "HIT 50%+":
                    cell.fill = green
                elif "THIN" in str(v) or "BARELY" in str(v):
                    cell.fill = yellow

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows)+1}"
    ws.freeze_panes = "A2"
    widths = [36, 22, 14, 16, 16, 42, 14, 16, 16, 14, 16, 14, 14, 14, 14, 16, 22, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32

    info = wb.create_sheet("Supplier")
    facts = [
        ("Legal name", "Ningbo King Smart Trading Co., Ltd."),
        ("Alibaba store", "https://kingsmartplumbing.en.alibaba.com/"),
        ("Made-in-China", "https://kingsmartplumbing.en.made-in-china.com/"),
        ("Website", "https://www.king-smart.cn/"),
        ("Type", "Trading company — NOT a factory. Founded 2024-03-05. 5 employees."),
        ("Contact", "George Ma (ex EZ-FLO 10 yrs, RWC 2 yrs)"),
        ("Email", "george.ma@king-smart.cn"),
        ("China office", "+86-574-86867745-101"),
        ("China cell", "+86-15158319134 / +86-18958232739"),
        ("US phone", "(509) 931-2251  (KING SMART Plumbing Master Co., Ltd., Washington)"),
        ("China addr", "527 Baoshan Road, Beilun, Ningbo, Zhejiang"),
        ("US stock story", "They pre-import into a WA warehouse + HK sister Climber Trading for logistics. 24h ship / ~7 day US delivery on in-stock SKUs. That is real — and it is why prices look like US wholesale, not factory FOB."),
        ("NSF", "King Smart is not the NSF listee. They claim NSF/UPC on listings; certificates come from whichever factory they buy from. Demand the actual NSF file + factory name before any PO."),
        ("Duty assumption", "China FOB × 1.428 (5.3% MFN + 25% 301 + 12.5% FLIP) + $0.03/pc freight"),
        ("USA stock", "Treat listed HIGH price as delivered (duty already inside). Do not add duty again."),
    ]
    info["A1"] = "Field"
    info["B1"] = "Value"
    info["A1"].font = header_font
    info["B1"].font = header_font
    info["A1"].fill = header_fill
    info["B1"].fill = header_fill
    for i, (k, v) in enumerate(facts, 2):
        info.cell(i, 1, k)
        info.cell(i, 2, v)
        info.cell(i, 2).alignment = Alignment(wrap_text=True)
    info.column_dimensions["A"].width = 22
    info.column_dimensions["B"].width = 110

    wb.save(path)


def main():
    rows = build_rows()
    write_csv(rows, OUT / "KingSmart_vs_NJPD.csv")
    write_xlsx(rows, OUT / "KingSmart_vs_NJPD.xlsx")
    print(f"Wrote {len(rows)} rows")
    for r in rows:
        print(
            f"{r['Item'][:32]:32} sell={r['NJPD_Sell_USD']:.4f} "
            f"usa={r['KingSmart_USA_stock_high_USD']} "
            f"m%={r['Margin_on_USA_stock_pct']}  {r['Verdict_at_USA_stock']}"
        )


if __name__ == "__main__":
    main()
